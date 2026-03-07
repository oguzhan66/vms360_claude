"""Reports Router - Uses Local Data Warehouse"""
from fastapi import APIRouter, Query, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from io import BytesIO
import json

from database import db
from auth import require_auth
from permissions import get_user_allowed_stores
from data_collector import (
    get_store_daily_data,
    get_store_hourly_data,
    get_date_range_data,
    get_all_stores_daily_data,
    get_latest_snapshots
)

router = APIRouter(prefix="/reports", tags=["Reports"])


# ============== HELPER FUNCTIONS ==============

async def get_filtered_store_ids(
    user: dict,
    store_ids: Optional[str] = None,
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None
) -> Optional[List[str]]:
    """Get store IDs filtered by user permissions and location filters"""
    allowed_stores = await get_user_allowed_stores(user)
    
    # Start with requested store IDs or all allowed
    if store_ids:
        requested_ids = store_ids.split(",")
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        filtered_ids = requested_ids
    elif allowed_stores is not None:
        if not allowed_stores:
            return []  # No access
        filtered_ids = list(allowed_stores)
    else:
        filtered_ids = None  # All stores
    
    # Apply location filters
    if region_id or city_id or district_id:
        # Build district filter
        district_ids = []
        if district_id:
            district_ids = [district_id]
        elif city_id:
            districts = await db.districts.find({"city_id": city_id}, {"_id": 0, "id": 1}).to_list(100)
            district_ids = [d["id"] for d in districts]
        elif region_id:
            cities = await db.cities.find({"region_id": region_id}, {"_id": 0, "id": 1}).to_list(100)
            city_ids = [c["id"] for c in cities]
            districts = await db.districts.find({"city_id": {"$in": city_ids}}, {"_id": 0, "id": 1}).to_list(500)
            district_ids = [d["id"] for d in districts]
        
        if district_ids:
            location_stores = await db.stores.find(
                {"district_id": {"$in": district_ids}},
                {"_id": 0, "id": 1}
            ).to_list(500)
            location_store_ids = {s["id"] for s in location_stores}
            
            if filtered_ids is not None:
                filtered_ids = [sid for sid in filtered_ids if sid in location_store_ids]
            else:
                filtered_ids = list(location_store_ids)
    
    return filtered_ids


def get_date_range_from_string(date_range: str, date_from: Optional[str], date_to: Optional[str]):
    """Convert date range string to actual dates"""
    now = datetime.now(timezone.utc)
    
    if date_from and date_to:
        return date_from, date_to
    
    if date_range == "1d":
        start = now.strftime("%Y-%m-%d")
        end = now.strftime("%Y-%m-%d")
    elif date_range == "1w":
        start = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        end = now.strftime("%Y-%m-%d")
    elif date_range == "1m":
        start = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        end = now.strftime("%Y-%m-%d")
    elif date_range == "1y":
        start = (now - timedelta(days=365)).strftime("%Y-%m-%d")
        end = now.strftime("%Y-%m-%d")
    else:
        start = now.strftime("%Y-%m-%d")
        end = now.strftime("%Y-%m-%d")
    
    return start, end


# ============== SUMMARY REPORT ==============

@router.get("/summary")
async def get_summary_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get summary report from local data warehouse"""
    filtered_store_ids = await get_filtered_store_ids(user, store_ids, region_id, city_id, district_id)
    
    if filtered_store_ids is not None and not filtered_store_ids:
        return {
            "counter_summary": {"total_stores": 0, "total_visitors": 0, "total_in": 0, "total_out": 0},
            "queue_summary": {"total_stores": 0, "total_queue_length": 0},
            "analytics_summary": {"total_events": 0},
            "stores": [],
            "queues": [],
            "data_source": "local_warehouse"
        }
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get daily summaries from data warehouse
    daily_data = await get_all_stores_daily_data(today, filtered_store_ids)
    
    # If no data in warehouse, get from latest snapshots
    if not daily_data:
        # Get latest counter snapshots
        query = {"date": today}
        if filtered_store_ids:
            query["store_id"] = {"$in": filtered_store_ids}
        
        counter_snapshots = await db.counter_snapshots.find(query, {"_id": 0}).to_list(500)
        queue_snapshots = await db.queue_snapshots.find(query, {"_id": 0}).to_list(500)
        
        stores_data = []
        for snap in counter_snapshots:
            stores_data.append({
                "store_id": snap.get("store_id"),
                "store_name": snap.get("store_name"),
                "total_in": snap.get("total_in", 0),
                "total_out": snap.get("total_out", 0),
                "current_visitors": snap.get("current_visitors", 0),
                "capacity": snap.get("capacity", 100),
                "occupancy_percent": snap.get("occupancy_percent", 0),
                "status": "critical" if snap.get("occupancy_percent", 0) >= 95 else "warning" if snap.get("occupancy_percent", 0) >= 80 else "normal"
            })
        
        queues_data = []
        for snap in queue_snapshots:
            queues_data.append({
                "store_id": snap.get("store_id"),
                "store_name": snap.get("store_name"),
                "total_queue_length": snap.get("total_queue_length", 0),
                "avg_wait_time_seconds": snap.get("avg_wait_time_seconds", 0)
            })
        
        return {
            "counter_summary": {
                "total_stores": len(stores_data),
                "total_visitors": sum(s["current_visitors"] for s in stores_data),
                "total_in": sum(s["total_in"] for s in stores_data),
                "total_out": sum(s["total_out"] for s in stores_data),
                "stores_critical": len([s for s in stores_data if s["status"] == "critical"]),
                "stores_warning": len([s for s in stores_data if s["status"] == "warning"]),
                "stores_normal": len([s for s in stores_data if s["status"] == "normal"])
            },
            "queue_summary": {
                "total_stores": len(queues_data),
                "total_queue_length": sum(q["total_queue_length"] for q in queues_data)
            },
            "analytics_summary": {"total_events": 0},
            "stores": stores_data,
            "queues": queues_data,
            "data_source": "local_warehouse_snapshots"
        }
    
    # Build response from daily summaries
    stores_data = []
    for d in daily_data:
        current = max(0, d.get("total_in", 0) - d.get("total_out", 0))
        occ = d.get("avg_occupancy", 0)
        stores_data.append({
            "store_id": d.get("store_id"),
            "store_name": d.get("store_name"),
            "district_name": d.get("district_name", ""),
            "city_name": d.get("city_name", ""),
            "region_name": d.get("region_name", ""),
            "total_in": d.get("total_in", 0),
            "total_out": d.get("total_out", 0),
            "current_visitors": current,
            "capacity": d.get("capacity", 100),
            "occupancy_percent": occ,
            "status": "critical" if occ >= 95 else "warning" if occ >= 80 else "normal"
        })
    
    return {
        "counter_summary": {
            "total_stores": len(stores_data),
            "total_visitors": sum(s["current_visitors"] for s in stores_data),
            "total_in": sum(s["total_in"] for s in stores_data),
            "total_out": sum(s["total_out"] for s in stores_data),
            "stores_critical": len([s for s in stores_data if s["status"] == "critical"]),
            "stores_warning": len([s for s in stores_data if s["status"] == "warning"]),
            "stores_normal": len([s for s in stores_data if s["status"] == "normal"])
        },
        "queue_summary": {
            "total_stores": len(daily_data),
            "total_queue_length": sum(d.get("max_queue_length", 0) for d in daily_data),
            "avg_wait_time_min": sum(d.get("avg_wait_time_min", 0) for d in daily_data) / len(daily_data) if daily_data else 0
        },
        "analytics_summary": {
            "total_events": sum(d.get("total_analytics_events", 0) for d in daily_data),
            "male_count": sum(d.get("male_count", 0) for d in daily_data),
            "female_count": sum(d.get("female_count", 0) for d in daily_data)
        },
        "stores": stores_data,
        "data_source": "local_warehouse"
    }


# ============== COUNTER REPORT ==============

@router.get("/counter")
async def get_counter_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get people counter report from local data warehouse"""
    filtered_store_ids = await get_filtered_store_ids(user, store_ids, region_id, city_id, district_id)
    start_date, end_date = get_date_range_from_string(date_range, date_from, date_to)
    
    if filtered_store_ids is not None and not filtered_store_ids:
        return {
            "report_type": "counter",
            "date_range": date_range,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {"total_stores": 0},
            "stores": [],
            "data_source": "local_warehouse"
        }
    
    # --- DOĞRU COUNTER HESAPLAMASI ---
    # VMS sayacı kümülatif çalışır: gün içinde artar, gece sıfırlanır.
    # Her günün ziyaretçi sayısı = o günün EN SON snapshot'ındaki total_in değeridir.
    # Çok günlük rapor için: her store için her günün son değerini al, sonra topla.

    match_stage = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        match_stage["store_id"] = {"$in": filtered_store_ids}

    # Her store+gün kombinasyonu için son snapshot
    daily_last_pipeline = [
        {"$match": match_stage},
        {"$sort": {"hour": -1, "minute": -1}},
        {"$group": {
            "_id": {"store_id": "$store_id", "date": "$date"},
            "total_in": {"$first": "$total_in"},
            "total_out": {"$first": "$total_out"},
            "store_name": {"$first": "$store_name"},
        }}
    ]
    snapshot_daily = await db.counter_snapshots.aggregate(daily_last_pipeline).to_list(50000)

    # Lokasyon bilgilerini daily_summaries'den al (varsa)
    location_map = {}
    loc_query = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        loc_query["store_id"] = {"$in": filtered_store_ids}
    for d in await db.daily_summaries.find(loc_query, {"_id": 0, "store_id": 1,
            "district_name": 1, "city_name": 1, "region_name": 1}).to_list(5000):
        sid = d.get("store_id")
        if sid and sid not in location_map:
            location_map[sid] = {
                "district_name": d.get("district_name", ""),
                "city_name": d.get("city_name", ""),
                "region_name": d.get("region_name", ""),
            }

    # Her store için günlük son değerleri topla
    store_totals = {}
    for row in snapshot_daily:
        sid = row["_id"]["store_id"]
        loc = location_map.get(sid, {"district_name": "", "city_name": "", "region_name": ""})
        if sid not in store_totals:
            store_totals[sid] = {
                "store_id": sid,
                "store_name": row.get("store_name", ""),
                "district_name": loc["district_name"],
                "city_name": loc["city_name"],
                "region_name": loc["region_name"],
                "total_in": 0,
                "total_out": 0,
                "days": 0
            }
        store_totals[sid]["total_in"] += row.get("total_in", 0)
        store_totals[sid]["total_out"] += row.get("total_out", 0)
        store_totals[sid]["days"] += 1

    stores = list(store_totals.values())
    for s in stores:
        s["current_visitors"] = max(0, s["total_in"] - s["total_out"])
        s["avg_daily_visitors"] = round(s["total_in"] / s["days"], 1) if s["days"] > 0 else 0

    return {
        "report_type": "counter",
        "date_range": date_range,
        "date_from": start_date,
        "date_to": end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_stores": len(stores),
            "total_in": sum(s["total_in"] for s in stores),
            "total_out": sum(s["total_out"] for s in stores),
            "current_visitors": sum(s["current_visitors"] for s in stores)
        },
        "stores": stores,
        "data_source": "counter_snapshots_daily_last"
    }


# ============== QUEUE REPORT ==============

@router.get("/queue")
async def get_queue_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get queue analysis report from local data warehouse"""
    filtered_store_ids = await get_filtered_store_ids(user, store_ids, region_id, city_id, district_id)
    start_date, end_date = get_date_range_from_string(date_range, date_from, date_to)
    
    if filtered_store_ids is not None and not filtered_store_ids:
        return {
            "report_type": "queue",
            "date_range": date_range,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {"total_stores": 0},
            "stores": [],
            "data_source": "local_warehouse"
        }
    
    # --- DOĞRU QUEUE HESAPLAMASI ---
    # queue_snapshots anlık ölçümler içerir.
    # Anlamlı metrikler: ortalama kuyruk, max kuyruk, eşik aşım sayısı.
    # daily_summaries güne ait özet içerir; eksik olabilir (gece oluşturulur).
    # Bu yüzden doğrudan queue_snapshots kullanıyoruz.

    match_stage = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        match_stage["store_id"] = {"$in": filtered_store_ids}

    # Stores bilgisi için - threshold ve lokasyon
    stores_info = await db.stores.find({}, {"_id": 0, "id": 1, "name": 1,
        "queue_threshold": 1, "district_id": 1}).to_list(500)
    store_threshold_map = {s["id"]: s.get("queue_threshold", 5) for s in stores_info}

    # Lokasyon bilgileri
    location_map = {}
    loc_query = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        loc_query["store_id"] = {"$in": filtered_store_ids}
    for d in await db.daily_summaries.find(loc_query, {"_id": 0, "store_id": 1,
            "store_name": 1, "district_name": 1, "city_name": 1, "region_name": 1}).to_list(5000):
        sid = d.get("store_id")
        if sid and sid not in location_map:
            location_map[sid] = {
                "store_name": d.get("store_name", ""),
                "district_name": d.get("district_name", ""),
                "city_name": d.get("city_name", ""),
                "region_name": d.get("region_name", ""),
            }

    # Her store için: ortalama kuyruk, max kuyruk, eşik aşım sayısı
    agg_pipeline = [
        {"$match": match_stage},
        {"$group": {
            "_id": "$store_id",
            "avg_queue": {"$avg": "$total_queue_length"},
            "max_queue": {"$max": "$total_queue_length"},
            "total_snapshots": {"$sum": 1},
            "store_name": {"$first": "$store_name"},
            "avg_wait_seconds": {"$avg": "$avg_wait_time_seconds"},
        }}
    ]
    agg_result = await db.queue_snapshots.aggregate(agg_pipeline).to_list(500)

    # Eşik aşım sayısı için ayrı sorgu (store_id bazında)
    stores = []
    for row in agg_result:
        sid = row["_id"]
        if not sid:
            continue  # store_id boş olanları atla (Bilinmiyor sorunu)
        threshold = store_threshold_map.get(sid, 5)
        loc = location_map.get(sid, {
            "store_name": row.get("store_name", ""),
            "district_name": "", "city_name": "", "region_name": ""
        })

        # Eşik aşım sayısı
        exceed_count = await db.queue_snapshots.count_documents({
            **match_stage, "store_id": sid,
            "total_queue_length": {"$gte": threshold}
        })

        stores.append({
            "store_id": sid,
            "store_name": loc.get("store_name") or row.get("store_name", ""),
            "district_name": loc["district_name"],
            "city_name": loc["city_name"],
            "region_name": loc["region_name"],
            "avg_queue_length": round(row.get("avg_queue", 0), 1),
            "max_queue_length": row.get("max_queue", 0),
            "avg_wait_time_min": round(row.get("avg_wait_seconds", 0) / 60, 1),
            "queue_threshold": threshold,
            "threshold_exceed_count": exceed_count,
            "total_snapshots": row.get("total_snapshots", 0),
        })

    return {
        "report_type": "queue",
        "date_range": date_range,
        "date_from": start_date,
        "date_to": end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_stores": len(stores),
            "avg_queue_length": round(sum(s["avg_queue_length"] for s in stores) / len(stores), 1) if stores else 0,
            "max_queue_length": max((s["max_queue_length"] for s in stores), default=0),
            "avg_wait_time_min": round(sum(s["avg_wait_time_min"] for s in stores) / len(stores), 1) if stores else 0
        },
        "stores": stores,
        "data_source": "queue_snapshots_aggregated"
    }


# ============== ANALYTICS REPORT ==============

@router.get("/analytics")
async def get_analytics_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get age/gender analytics report from local data warehouse"""
    filtered_store_ids = await get_filtered_store_ids(user, store_ids, region_id, city_id, district_id)
    start_date, end_date = get_date_range_from_string(date_range, date_from, date_to)
    
    if filtered_store_ids is not None and not filtered_store_ids:
        return {
            "report_type": "analytics",
            "date_range": date_range,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "summary": {"total_detections": 0},
            "gender_distribution": {},
            "age_distribution": {},
            "data_source": "local_warehouse"
        }
    
    # --- DOĞRU ANALYTICS HESAPLAMASI ---
    # daily_summaries gece oluşturulur; gün içinde veya tarih aralığı için boş olabilir.
    # Doğrudan analytics_snapshots kullanıyoruz - her snapshot 5dk'lık veri içerir.
    # Her snapshot'ın gender ve age verilerini topluyoruz.

    analytics_match = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        analytics_match["store_id"] = {"$in": filtered_store_ids}

    # Lokasyon bilgileri
    location_map = {}
    loc_query = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_store_ids:
        loc_query["store_id"] = {"$in": filtered_store_ids}
    for d in await db.daily_summaries.find(loc_query, {"_id": 0, "store_id": 1,
            "store_name": 1, "district_name": 1, "city_name": 1, "region_name": 1}).to_list(5000):
        sid = d.get("store_id")
        if sid and sid not in location_map:
            location_map[sid] = {
                "store_name": d.get("store_name", ""),
                "district_name": d.get("district_name", ""),
                "city_name": d.get("city_name", ""),
                "region_name": d.get("region_name", ""),
            }

    analytics_data = await db.analytics_snapshots.find(analytics_match, {"_id": 0}).to_list(100000)

    # Toplam cinsiyet ve yaş dağılımı
    total_male = 0
    total_female = 0
    total_unknown = 0
    age_totals = {"0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0}

    # Store bazında toplama
    store_agg = {}
    for a in analytics_data:
        sid = a.get("store_id")
        if not sid:
            continue
        gd = a.get("gender_distribution", {})
        ad = a.get("age_distribution", {})
        m = gd.get("Male", 0)
        f = gd.get("Female", 0)
        u = gd.get("Unknown", 0)
        total_male += m
        total_female += f
        total_unknown += u
        for age_group, count in ad.items():
            if age_group in age_totals:
                age_totals[age_group] += count
        if sid not in store_agg:
            loc = location_map.get(sid, {"store_name": a.get("store_name",""),
                "district_name":"","city_name":"","region_name":""})
            store_agg[sid] = {
                "store_id": sid,
                "store_name": loc.get("store_name") or a.get("store_name", ""),
                "district_name": loc["district_name"],
                "city_name": loc["city_name"],
                "region_name": loc["region_name"],
                "male_count": 0, "female_count": 0,
                "total_events": 0, "snapshot_count": 0
            }
        store_agg[sid]["male_count"] += m
        store_agg[sid]["female_count"] += f
        store_agg[sid]["total_events"] += a.get("total_events", 0)
        store_agg[sid]["snapshot_count"] += 1

    stores = list(store_agg.values())
    total = total_male + total_female + total_unknown

    return {
        "report_type": "analytics",
        "date_range": date_range,
        "date_from": start_date,
        "date_to": end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_detections": total,
            "male_count": total_male,
            "female_count": total_female,
            "unknown_count": total_unknown,
            "male_percent": round(total_male / total * 100, 1) if total > 0 else 0,
            "female_percent": round(total_female / total * 100, 1) if total > 0 else 0
        },
        "gender_distribution": {
            "Male": total_male,
            "Female": total_female,
            "Unknown": total_unknown
        },
        "age_distribution": age_totals,
        "stores": stores,
        "data_source": "analytics_snapshots_direct"
    }


# ============== EXPORT ENDPOINT ==============

@router.get("/export")
async def export_report(
    report_type: str = "counter",
    format: str = "json",
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Export report data in various formats"""
    # Get report data based on type
    if report_type == "counter":
        data = await get_counter_report(
            region_id, city_id, district_id, store_ids,
            date_range, date_from, date_to, user
        )
    elif report_type == "queue":
        data = await get_queue_report(
            region_id, city_id, district_id, store_ids,
            date_range, date_from, date_to, user
        )
    elif report_type == "analytics":
        data = await get_analytics_report(
            region_id, city_id, district_id, store_ids,
            date_range, date_from, date_to, user
        )
    else:
        data = await get_summary_report(
            region_id, city_id, district_id, store_ids, user
        )
    
    if format == "json":
        return data
    elif format == "csv":
        # Generate CSV
        import csv
        from io import StringIO
        
        output = StringIO()
        stores = data.get("stores", [])
        if stores:
            writer = csv.DictWriter(output, fieldnames=stores[0].keys())
            writer.writeheader()
            writer.writerows(stores)
        
        content = output.getvalue()
        return StreamingResponse(
            BytesIO(content.encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.csv"}
        )
    elif format == "excel":
        # Generate Excel
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            stores = data.get("stores", [])
            if stores:
                # Write headers
                headers = list(stores[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Write data
                for row_idx, store in enumerate(stores, 2):
                    for col_idx, key in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=store.get(key, ""))
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel export not available")
    
    return data
