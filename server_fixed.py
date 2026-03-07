"""
VMS360 Retail Panel - Backend Server
Modular architecture with FastAPI routers
"""
from fastapi import FastAPI, APIRouter, HTTPException, Query, BackgroundTasks, Depends, Request
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import asyncio
import xml.etree.ElementTree as ET
from io import BytesIO
import json
from jose import JWTError, jwt
from passlib.context import CryptContext
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ============== REGISTER TURKISH FONTS FOR PDF ==============
# DejaVu Sans has full Turkish character support (ğ, ş, ı, ç, ü, ö, İ, Ğ, Ş, Ç, Ü, Ö)
try:
    pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuBold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    PDF_FONT = 'DejaVu'
    PDF_FONT_BOLD = 'DejaVuBold'
except:
    try:
        pdfmetrics.registerFont(TTFont('Liberation', '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('LiberationBold', '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'))
        PDF_FONT = 'Liberation'
        PDF_FONT_BOLD = 'LiberationBold'
    except:
        PDF_FONT = 'Helvetica'
        PDF_FONT_BOLD = 'Helvetica-Bold'

# ============== SENTRY INITIALIZATION (P1) ==============
import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.starlette import StarletteIntegration
from sentry_sdk.integrations.logging import LoggingIntegration

SENTRY_DSN = os.environ.get("SENTRY_DSN")
if SENTRY_DSN:
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        environment=os.environ.get("ENVIRONMENT", "development"),
        traces_sample_rate=0.1,  # 10% of transactions
        profiles_sample_rate=0.1,
        integrations=[
            StarletteIntegration(transaction_style="endpoint"),
            FastApiIntegration(transaction_style="endpoint"),
            LoggingIntegration(level=logging.INFO, event_level=logging.ERROR),
        ],
        send_default_pii=False,
        attach_stacktrace=True,
        release=os.environ.get("APP_VERSION", "1.0.0"),
    )
    logging.info("Sentry initialized successfully")

# Import from modular structure
from database import db, client
from vms_utils import fetch_vms_data as fetch_vms_data_util, parse_counter_xml, parse_queue_xml, parse_analytics_xml, parse_camera_list_xml

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Import and include all routers
from routers.analytics import router as analytics_router
from routers.auth import router as auth_router, users_router
from routers.vms import router as vms_router
from routers.locations import router as locations_router
from routers.stores import router as stores_router
from routers.cameras import router as cameras_router
from routers.settings import router as settings_router
from routers.live import router as live_router
from routers.historical import router as historical_router
from routers.scheduled_reports import router as scheduled_reports_router
from routers.local_data import router as local_data_router
from routers.floors import router as floors_router
from routers.heatmap import router as heatmap_router

# Include routers - LOCAL DATA ROUTER FIRST (takes precedence for live/reports endpoints)
# All data now comes from local MongoDB warehouse, not live VMS
api_router.include_router(local_data_router)  # MUST BE FIRST - overrides live/reports
api_router.include_router(analytics_router)
api_router.include_router(auth_router)
api_router.include_router(users_router)
api_router.include_router(locations_router)
api_router.include_router(stores_router)
api_router.include_router(cameras_router)
api_router.include_router(settings_router)
api_router.include_router(historical_router)
# DISABLED: scheduled_reports_router - Using server.py endpoints instead (they have date_range support)
# api_router.include_router(scheduled_reports_router)
api_router.include_router(floors_router)
api_router.include_router(heatmap_router)
# Note: VMS is now only used by data_collector.py for background data collection

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============== AUTH CONFIG ==============
SECRET_KEY = os.environ.get("SECRET_KEY", "vms360-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if credentials is None:
        return None
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None:
            return None
        return {"username": username, "role": role}
    except JWTError:
        return None

async def require_auth(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if credentials is None:
        raise HTTPException(status_code=401, detail="Giriş gerekli")
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None:
            raise HTTPException(status_code=401, detail="Geçersiz token")
        return {"username": username, "role": role}
    except JWTError:
        raise HTTPException(status_code=401, detail="Geçersiz token")

async def require_admin(user: dict = Depends(require_auth)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin yetkisi gerekli")
    return user

# ============== MODELS ==============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    full_name: str
    role: str = "operator"  # admin, operator
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    role: str = "operator"
    allowed_region_ids: List[str] = Field(default_factory=list)
    allowed_city_ids: List[str] = Field(default_factory=list)
    allowed_store_ids: List[str] = Field(default_factory=list)

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    full_name: str
    role: str
    is_active: bool
    allowed_region_ids: List[str] = Field(default_factory=list)
    allowed_city_ids: List[str] = Field(default_factory=list)
    allowed_store_ids: List[str] = Field(default_factory=list)

class VMSServer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    url: str
    username: str
    password: Optional[str] = ""
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VMSServerCreate(BaseModel):
    name: str
    url: str
    username: str
    password: Optional[str] = ""

class VMSServerUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    is_active: Optional[bool] = None

class Region(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class City(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    region_id: str
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class District(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    city_id: str
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Store(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    district_id: str
    vms_id: str
    capacity: int = 100
    queue_threshold: int = 5
    # Camera assignments - stores VMS camera IDs
    counter_camera_id: Optional[str] = None  # Selected counter camera
    queue_camera_id: Optional[str] = None    # Selected queue camera
    analytics_camera_id: Optional[str] = None # Selected analytics camera
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Camera(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    store_id: str
    camera_vms_id: str  # Camera ID from VMS system
    name: str
    type: str = "counter"  # counter, queue, analytics
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============== SMTP & SCHEDULED REPORTS MODELS ==============

class SMTPSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    host: str
    port: int = 587
    username: str
    password: str
    from_email: str
    from_name: str = "VMS360 Rapor Sistemi"
    use_tls: bool = True
    is_active: bool = True
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SMTPSettingsCreate(BaseModel):
    host: str
    port: int = 587
    username: str
    password: str
    from_email: str
    from_name: str = "VMS360 Rapor Sistemi"
    use_tls: bool = True

class SMTPTestRequest(BaseModel):
    test_email: str

class ScheduledReport(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    report_type: str  # hourly_traffic, weekday_comparison, store_comparison, queue_analysis, demographics, all
    format: str = "excel"  # excel, csv, json
    frequency: str = "daily"  # daily, weekly, monthly
    send_time: str = "08:00"  # HH:MM format
    send_day: Optional[int] = None  # For weekly: 0-6 (Mon-Sun), For monthly: 1-31
    recipients: List[str] = []
    is_active: bool = True
    last_sent: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # New filter fields
    store_ids: List[str] = []  # Empty = all stores
    date_range: str = "1d"  # 1d, 7d, 30d, custom
    date_from: Optional[str] = None  # For custom range: YYYY-MM-DD
    date_to: Optional[str] = None  # For custom range: YYYY-MM-DD
    # Report-specific filters
    hour_from: Optional[int] = None  # For hourly: start hour (0-23)
    hour_to: Optional[int] = None  # For hourly: end hour (0-23)
    gender_filter: Optional[str] = None  # For demographics: Male, Female, or None (all)
    min_queue_length: Optional[int] = None  # For queue: minimum queue to include

class ScheduledReportCreate(BaseModel):
    name: str
    report_type: str
    format: str = "excel"
    frequency: str = "daily"
    send_time: str = "08:00"
    send_day: Optional[int] = None
    recipients: List[str]
    # New filter fields
    store_ids: List[str] = []
    date_range: str = "1d"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    hour_from: Optional[int] = None
    hour_to: Optional[int] = None
    gender_filter: Optional[str] = None
    min_queue_length: Optional[int] = None

class ScheduledReportUpdate(BaseModel):
    name: Optional[str] = None
    report_type: Optional[str] = None
    format: Optional[str] = None
    frequency: Optional[str] = None
    send_time: Optional[str] = None
    send_day: Optional[int] = None
    recipients: Optional[List[str]] = None
    is_active: Optional[bool] = None
    # New filter fields
    store_ids: Optional[List[str]] = None
    date_range: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    hour_from: Optional[int] = None
    hour_to: Optional[int] = None
    gender_filter: Optional[str] = None
    min_queue_length: Optional[int] = None

# ============== HISTORICAL DATA MODELS ==============

class HistoricalCounter(BaseModel):
    """Stores historical counter data for reporting"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    store_id: str
    store_name: str
    date: str  # YYYY-MM-DD format
    hour: int  # 0-23
    total_in: int = 0
    total_out: int = 0
    current_visitors: int = 0
    occupancy_percent: float = 0
    status: str = "normal"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HistoricalQueue(BaseModel):
    """Stores historical queue data for reporting"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    store_id: str
    store_name: str
    date: str
    hour: int
    total_queue_length: int = 0
    avg_wait_minutes: float = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HistoricalAnalytics(BaseModel):
    """Stores historical demographics data"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    hour: int
    total_detections: int = 0
    male_count: int = 0
    female_count: int = 0
    age_0_17: int = 0
    age_18_24: int = 0
    age_25_34: int = 0
    age_35_44: int = 0
    age_45_54: int = 0
    age_55_plus: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StoreCreate(BaseModel):
    name: str
    district_id: str
    vms_id: str
    capacity: int = 100
    queue_threshold: int = 5
    counter_camera_id: Optional[str] = None
    queue_camera_id: Optional[str] = None
    analytics_camera_id: Optional[str] = None

class StoreUpdate(BaseModel):
    name: Optional[str] = None
    district_id: Optional[str] = None
    vms_id: Optional[str] = None
    capacity: Optional[int] = None
    queue_threshold: Optional[int] = None
    counter_camera_id: Optional[str] = None
    queue_camera_id: Optional[str] = None
    analytics_camera_id: Optional[str] = None

class CameraCreate(BaseModel):
    store_id: str
    camera_vms_id: str
    name: str
    type: str = "counter"

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "global_settings"
    refresh_interval: int = 30  # seconds
    capacity_warning_percent: int = 80
    capacity_critical_percent: int = 95
    email_notifications: bool = False
    notification_email: Optional[str] = None

class LocationCreate(BaseModel):
    name: str
    parent_id: Optional[str] = None

# ============== HELPER FUNCTIONS ==============

async def fetch_vms_data(vms: dict, endpoint: str) -> Optional[str]:
    """Fetch data from VMS server"""
    try:
        url = f"{vms['url']}{endpoint}"
        auth = None
        if vms.get('username'):
            auth = (vms['username'], vms.get('password', ''))
        
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(url, auth=auth if auth else None)
            if response.status_code == 200:
                return response.text
            logger.warning(f"VMS request failed: {response.status_code}")
            return None
    except Exception as e:
        logger.error(f"VMS connection error: {e}")
        return None

def parse_counter_xml(xml_string: str) -> Dict:
    """Parse people counter XML response - Returns dict with 'cameras' key"""
    try:
        root = ET.fromstring(xml_string)
        cameras = []
        for camera_state in root.findall('.//CameraState'):
            cameras.append({
                'camera_id': camera_state.findtext('CameraID', ''),
                'camera_name': camera_state.findtext('CameraName', ''),
                'last_reset': camera_state.findtext('LastResetTime', ''),
                'counters': [{
                    'index': '0',
                    'in_count': int(camera_state.findtext('In', '0')),
                    'out_count': int(camera_state.findtext('Out', '0'))
                }]
            })
        return {'cameras': cameras}
    except Exception as e:
        logger.error(f"XML parse error: {e}")
        return {'cameras': []}

def parse_queue_xml(xml_string: str) -> Dict:
    """Parse queue detection XML response - Returns dict with 'cameras' key"""
    try:
        root = ET.fromstring(xml_string)
        cameras = []
        for camera_state in root.findall('.//CameraState'):
            camera_id = camera_state.findtext('CameraID', '')
            camera_name = camera_state.findtext('CameraName', '')
            zones = []
            for zone_state in camera_state.findall('.//ZoneState'):
                zones.append({
                    'zone_index': int(zone_state.findtext('ZoneIndex', '0')),
                    'queue_length': int(zone_state.findtext('QueueLength', '0')),
                    'is_queue': zone_state.findtext('IsQueue', 'false').lower() == 'true'
                })
            cameras.append({
                'camera_id': camera_id,
                'camera_name': camera_name,
                'zones': zones
            })
        return {'cameras': cameras}
    except Exception as e:
        logger.error(f"XML parse error: {e}")
        return {'cameras': []}

def parse_analytics_xml(xml_string: str) -> Dict:
    """Parse face recognition/analytics XML response - Returns dict with 'cameras' key"""
    try:
        root = ET.fromstring(xml_string)
        detections = []
        for item in root.findall('.//Item'):
            detections.append({
                'event_id': item.findtext('EventID', ''),
                'time': item.findtext('Time', ''),
                'camera_id': item.findtext('CameraID', ''),
                'camera_name': item.findtext('CameraName', ''),
                'age': int(item.findtext('Age', '0')),
                'gender': item.findtext('Gender', 'Unknown'),
                'is_recognized': item.findtext('IsRecognized', 'False') == 'True'
            })
        # Group by camera
        cameras_dict = {}
        for det in detections:
            cam_id = det['camera_id']
            if cam_id not in cameras_dict:
                cameras_dict[cam_id] = {
                    'camera_id': cam_id,
                    'camera_name': det.get('camera_name', ''),
                    'detections': []
                }
            cameras_dict[cam_id]['detections'].append(det)
        return {'cameras': list(cameras_dict.values())}
    except Exception as e:
        logger.error(f"XML parse error: {e}")
        return {'cameras': []}

# ============== AUTH ENDPOINTS (Now handled by routers/auth.py) ==============
# Auth endpoints moved to routers/auth.py

# Initialize default users
async def init_default_users():
    """Create default admin and operator users if not exists"""
    # Create admin user
    existing_admin = await db.users.find_one({"username": "admin"})
    if not existing_admin:
        admin = User(
            username="admin",
            password_hash=get_password_hash("12345"),
            full_name="Sistem Yöneticisi",
            role="admin"
        )
        doc = admin.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Default admin user created: admin / 12345")
    
    # Create operator user
    existing_operator = await db.users.find_one({"username": "operator"})
    if not existing_operator:
        operator = User(
            username="operator",
            password_hash=get_password_hash("12345"),
            full_name="Operatör Kullanıcı",
            role="operator"
        )
        doc = operator.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Default operator user created: operator / 12345")

@app.on_event("startup")
async def startup_event():
    await init_default_users()
    await init_redis_cache()
    await init_scheduler()


async def init_redis_cache():
    """Initialize Redis cache connection"""
    from cache import init_redis
    await init_redis()


# ============== SCHEDULER FOR DATA COLLECTION & REPORTS ==============
scheduler = AsyncIOScheduler()

async def init_scheduler():
    """Initialize the scheduler and add jobs for data collection and scheduled reports"""
    from data_collector import (
        collect_all_snapshots, 
        create_hourly_aggregates, 
        create_daily_summary,
        cleanup_old_snapshots,
        check_store_health
    )
    from database import create_indexes
    
    # Create database indexes (P0: MongoDB Index)
    await create_indexes()
    
    # Job 1: Collect snapshots every 5 minutes
    scheduler.add_job(
        collect_all_snapshots,
        CronTrigger(minute='*/5'),  # Run every 5 minutes
        id='collect_snapshots',
        replace_existing=True
    )
    
    # Job 2: Create hourly aggregates at the end of each hour
    scheduler.add_job(
        create_hourly_aggregates,
        CronTrigger(minute='55'),  # Run at :55 of every hour
        id='create_hourly_aggregates',
        replace_existing=True
    )
    
    # Job 3: Create daily summary at end of day (23:59)
    scheduler.add_job(
        create_daily_summary,
        CronTrigger(hour='23', minute='59'),
        id='create_daily_summary',
        replace_existing=True
    )
    
    # Job 4: Cleanup old snapshots weekly (keep 7 days of snapshots)
    scheduler.add_job(
        cleanup_old_snapshots,
        CronTrigger(day_of_week='sun', hour='3', minute='0'),  # Sunday 3:00 AM
        id='cleanup_old_snapshots',
        replace_existing=True
    )
    
    # Job 5: Check scheduled reports every minute
    scheduler.add_job(
        check_scheduled_reports,
        CronTrigger(minute='*'),
        id='check_scheduled_reports',
        replace_existing=True
    )
    
    # Job 6: Legacy hourly data collection (for compatibility)
    scheduler.add_job(
        collect_historical_data,
        CronTrigger(minute='0'),
        id='collect_historical_data',
        replace_existing=True
    )
    
    # Job 7: P0 - Check store health every 10 minutes (veri gelmeme alarmı)
    scheduler.add_job(
        check_store_health,
        CronTrigger(minute='*/10'),  # Run every 10 minutes
        id='check_store_health',
        replace_existing=True
    )
    
    scheduler.start()
    logger.info("Scheduler started:")
    logger.info("  - Data snapshots: every 5 minutes")
    logger.info("  - Hourly aggregates: every hour at :55")
    logger.info("  - Daily summaries: every day at 23:59")
    logger.info("  - Snapshot cleanup: Sundays at 03:00")
    logger.info("  - Scheduled reports: every minute")
    logger.info("  - Store health check: every 10 minutes")

async def collect_historical_data():
    """Collect and store historical data from VMS"""
    try:
        now = datetime.now(timezone.utc)
        date_str = now.strftime("%Y-%m-%d")
        hour = now.hour
        
        logger.info(f"Collecting historical data for {date_str} hour {hour}")
        
        # Get counter data
        counter_data = await _fetch_live_counter_data(None, None)
        for store in counter_data:
            # Check if already collected for this hour
            existing = await db.historical_counter.find_one({
                "store_id": store["store_id"],
                "date": date_str,
                "hour": hour
            })
            
            if not existing:
                hist = HistoricalCounter(
                    store_id=store["store_id"],
                    store_name=store["store_name"],
                    date=date_str,
                    hour=hour,
                    total_in=store["total_in"],
                    total_out=store["total_out"],
                    current_visitors=store["current_visitors"],
                    occupancy_percent=store["occupancy_percent"],
                    status=store["status"]
                )
                doc = hist.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.historical_counter.insert_one(doc)
        
        # Get queue data
        queue_data = await _fetch_live_queue_data(None, None)
        for store in queue_data:
            existing = await db.historical_queue.find_one({
                "store_id": store["store_id"],
                "date": date_str,
                "hour": hour
            })
            
            if not existing:
                hist = HistoricalQueue(
                    store_id=store["store_id"],
                    store_name=store["store_name"],
                    date=date_str,
                    hour=hour,
                    total_queue_length=store["total_queue_length"],
                    avg_wait_minutes=store["total_queue_length"] * 2  # Estimated
                )
                doc = hist.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.historical_queue.insert_one(doc)
        
        # Get analytics data
        analytics = await get_live_analytics_data(None)
        existing = await db.historical_analytics.find_one({
            "date": date_str,
            "hour": hour
        })
        
        if not existing and analytics["total_events"] > 0:
            gender = analytics["gender_distribution"]
            age = analytics["age_distribution"]
            
            hist = HistoricalAnalytics(
                date=date_str,
                hour=hour,
                total_detections=analytics["total_events"],
                male_count=gender.get("Male", 0),
                female_count=gender.get("Female", 0),
                age_0_17=age.get("0-17", 0),
                age_18_24=age.get("18-24", 0),
                age_25_34=age.get("25-34", 0),
                age_35_44=age.get("35-44", 0),
                age_45_54=age.get("45-54", 0),
                age_55_plus=age.get("55+", 0)
            )
            doc = hist.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.historical_analytics.insert_one(doc)
        
        logger.info(f"Historical data collected successfully for {date_str} hour {hour}")
    
    except Exception as e:
        logger.error(f"Error collecting historical data: {str(e)}")

async def check_scheduled_reports():
    """Check if any scheduled reports need to be sent"""
    try:
        now = datetime.now(timezone.utc)
        current_time = now.strftime("%H:%M")
        current_weekday = now.weekday()  # 0=Monday, 6=Sunday
        current_day = now.day
        
        # Get all active scheduled reports
        reports = await db.scheduled_reports.find({"is_active": True}, {"_id": 0}).to_list(100)
        smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
        
        if not smtp_settings:
            return
        
        for report in reports:
            should_send = False
            send_time = report.get("send_time", "08:00")
            
            # Check if it's time to send based on frequency
            if current_time == send_time:
                frequency = report.get("frequency", "daily")
                
                if frequency == "daily":
                    should_send = True
                elif frequency == "weekly":
                    send_day = report.get("send_day", 0)
                    if current_weekday == send_day:
                        should_send = True
                elif frequency == "monthly":
                    send_day = report.get("send_day", 1)
                    if current_day == send_day:
                        should_send = True
                
                # Check if already sent today
                last_sent = report.get("last_sent")
                if last_sent:
                    if isinstance(last_sent, str):
                        last_sent = datetime.fromisoformat(last_sent.replace('Z', '+00:00'))
                    if last_sent.date() == now.date():
                        should_send = False  # Already sent today
            
            if should_send:
                logger.info(f"Sending scheduled report: {report['name']}")
                await send_scheduled_report(report, smtp_settings)
    
    except Exception as e:
        logger.error(f"Error checking scheduled reports: {str(e)}")

# ============== VMS ENDPOINTS ==============

@api_router.get("/")
async def root():
    return {"message": "Sagitech VMS Dashboard API"}


@api_router.get("/debug/data-check")
async def debug_data_check():
    """Debug endpoint to check data sources - Admin only"""
    from datetime import timedelta
    
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    week_ago = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    
    result = {
        "timestamp": now.isoformat(),
        "date_range": {"from": week_ago, "to": today},
        "tables": {}
    }
    
    # 1. Check counter_snapshots
    snap_total = await db.counter_snapshots.count_documents({})
    snap_today = await db.counter_snapshots.count_documents({"date": today})
    snap_week = await db.counter_snapshots.count_documents({"date": {"$gte": week_ago, "$lte": today}})
    
    # Get latest snapshot
    latest_snap = await db.counter_snapshots.find_one(
        {}, {"_id": 0}, sort=[("timestamp", -1)]
    )
    
    # Get sum of total_in from latest snapshots per store per day
    pipeline = [
        {"$match": {"date": {"$gte": week_ago, "$lte": today}}},
        {"$sort": {"timestamp": -1}},
        {"$group": {
            "_id": {"store_id": "$store_id", "date": "$date"},
            "latest": {"$first": "$$ROOT"}
        }},
        {"$group": {
            "_id": None,
            "total_in_sum": {"$sum": "$latest.total_in"},
            "total_out_sum": {"$sum": "$latest.total_out"},
            "count": {"$sum": 1}
        }}
    ]
    snap_agg = await db.counter_snapshots.aggregate(pipeline).to_list(1)
    
    result["tables"]["counter_snapshots"] = {
        "total_records": snap_total,
        "today_records": snap_today,
        "week_records": snap_week,
        "latest_snapshot": latest_snap,
        "week_aggregation": snap_agg[0] if snap_agg else None
    }
    
    # 2. Check daily_summaries
    summary_total = await db.daily_summaries.count_documents({})
    summary_week = await db.daily_summaries.count_documents({"date": {"$gte": week_ago, "$lte": today}})
    
    # Get sum from daily_summaries
    summaries = await db.daily_summaries.find(
        {"date": {"$gte": week_ago, "$lte": today}}, {"_id": 0}
    ).to_list(1000)
    
    summary_totals = {
        "total_in": sum(s.get("total_in", 0) for s in summaries),
        "total_out": sum(s.get("total_out", 0) for s in summaries)
    }
    
    result["tables"]["daily_summaries"] = {
        "total_records": summary_total,
        "week_records": summary_week,
        "week_totals": summary_totals
    }
    
    # 3. Check stores
    stores = await db.stores.find({}, {"_id": 0, "id": 1, "name": 1, "counter_camera_ids": 1, "counter_camera_id": 1}).to_list(100)
    result["tables"]["stores"] = {
        "count": len(stores),
        "list": stores[:10]  # First 10
    }
    
    # 4. Quick recommendations
    issues = []
    if snap_total == 0:
        issues.append("counter_snapshots tablosu BOŞ! VMS'den veri toplanmıyor olabilir.")
    if summary_total == 0:
        issues.append("daily_summaries tablosu BOŞ! Günlük özet job'ı çalışmamış olabilir.")
    if snap_week > 0 and (snap_agg and snap_agg[0]["total_in_sum"] == 0):
        issues.append("counter_snapshots var ama total_in değerleri 0!")
    
    result["issues"] = issues if issues else ["Tablo yapıları normal görünüyor"]
    
    return result


@api_router.get("/health")
async def get_health_status():
    """Get system health status - P0: Veri gelmeme alarmı"""
    from data_collector import get_system_health
    return await get_system_health()


@api_router.get("/health/stores")
async def get_stores_health(user: dict = Depends(require_auth)):
    """Get detailed store health status (requires auth)"""
    from data_collector import get_system_health
    health = await get_system_health()
    return health


@api_router.get("/debug/test-scheduled-report")
async def test_scheduled_report_generation(report_type: str = "analytics"):
    """Debug endpoint to test scheduled report data generation"""
    try:
        filters = {
            "store_ids": [],
            "date_range": "1d",
            "hour_from": None,
            "hour_to": None,
            "gender_filter": None,
            "min_queue_length": 0
        }
        
        report_data = await generate_report_data(report_type, filters)
        
        return {
            "success": True,
            "report_type": report_type,
            "data_type": report_data.get("type"),
            "data_count": len(report_data.get("data", [])),
            "data_sample": report_data.get("data", [])[:3] if report_data.get("data") else [],
            "full_data": report_data.get("data", [])
        }
    except Exception as e:
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@api_router.get("/debug/test-scheduled-excel")
async def test_scheduled_excel_generation(report_type: str = "analytics"):
    """Debug endpoint to test scheduled report Excel generation and download"""
    try:
        import xlsxwriter
        
        filters = {
            "store_ids": [],
            "date_range": "1d",
            "hour_from": None,
            "hour_to": None,
            "gender_filter": None,
            "min_queue_length": 0
        }
        
        report_data = await generate_report_data(report_type, filters)
        data = report_data.get("data", [])
        
        if not data:
            return {"success": False, "error": "No data returned from generate_report_data"}
        
        # Create Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        header_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#3B82F6', 
            'font_color': 'white', 
            'border': 1,
            'align': 'center'
        })
        cell_format = workbook.add_format({'border': 1, 'align': 'left'})
        number_format = workbook.add_format({'border': 1, 'num_format': '#,##0', 'align': 'right'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1})
        
        worksheet = workbook.add_worksheet("Rapor")
        
        if isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
                worksheet.set_column(col, col, 18)
            
            for row, item in enumerate(data, 1):
                is_total = item.get("Mağaza") == "TOPLAM"
                for col, key in enumerate(headers):
                    value = item.get(key, "")
                    if is_total:
                        worksheet.write(row, col, value, total_format)
                    elif isinstance(value, (int, float)):
                        worksheet.write(row, col, value, number_format)
                    else:
                        worksheet.write(row, col, value, cell_format)
        
        workbook.close()
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=test_rapor_{report_type}.xlsx"}
        )
    except Exception as e:
        import traceback
        return {"success": False, "error": str(e), "traceback": traceback.format_exc()}



@api_router.get("/debug/analytics-check")
async def debug_analytics_check():
    """Debug endpoint to check why analytics_snapshots might have 0 events"""
    from datetime import timedelta
    
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    result = {
        "timestamp": now.isoformat(),
        "stores_check": [],
        "vms_check": [],
        "analytics_snapshots_sample": [],
        "recommendations": []
    }
    
    # Check stores for analytics camera assignments
    stores = await db.stores.find({}, {"_id": 0}).to_list(100)
    for store in stores:
        analytics_camera_id = store.get("analytics_camera_id")
        analytics_camera_ids = store.get("analytics_camera_ids", [])
        result["stores_check"].append({
            "store_id": store["id"],
            "store_name": store["name"],
            "analytics_camera_id": analytics_camera_id,
            "analytics_camera_ids": analytics_camera_ids,
            "has_analytics_camera": bool(analytics_camera_id or analytics_camera_ids)
        })
    
    # Check if any store has analytics camera
    stores_with_analytics = [s for s in result["stores_check"] if s["has_analytics_camera"]]
    
    if not stores_with_analytics:
        result["recommendations"].append("❌ HİÇBİR MAĞAZADA ANALİTİK KAMERA TANIMLI DEĞİL! Mağaza ayarlarından analitik kamera atayın.")
    else:
        result["recommendations"].append(f"✅ {len(stores_with_analytics)} mağazada analitik kamera tanımlı")
    
    # Check VMS servers
    vms_servers = await db.vms_servers.find({}, {"_id": 0}).to_list(10)
    result["vms_check"] = [{"id": v["id"], "name": v["name"], "is_active": v.get("is_active", False)} for v in vms_servers]
    
    # Check sample analytics snapshots
    sample_snapshots = await db.analytics_snapshots.find(
        {"date": today},
        {"_id": 0, "store_id": 1, "store_name": 1, "total_events": 1, "gender_distribution": 1, "timestamp": 1}
    ).sort("timestamp", -1).limit(5).to_list(5)
    result["analytics_snapshots_sample"] = sample_snapshots
    
    # Check if snapshots exist but have 0 events
    snapshots_with_events = await db.analytics_snapshots.count_documents({
        "date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")},
        "total_events": {"$gt": 0}
    })
    total_snapshots = await db.analytics_snapshots.count_documents({
        "date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}
    })
    
    result["snapshots_with_events_last_week"] = snapshots_with_events
    result["total_snapshots_last_week"] = total_snapshots
    
    if total_snapshots > 0 and snapshots_with_events == 0:
        result["recommendations"].append("⚠️ Snapshot'lar kaydediliyor ama hiçbirinde event yok. VMS'den analytics verisi gelmemiş olabilir.")
        result["recommendations"].append("📝 VMS'deki yüz tanıma modülünün aktif olduğundan emin olun.")
    
    return result


@api_router.get("/debug/vms-analytics-raw")
async def debug_vms_analytics_raw():
    """Debug endpoint to see RAW VMS analytics data - to diagnose why snapshots are empty"""
    from vms_utils import fetch_vms_data, parse_analytics_xml
    
    result = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "vms_servers": [],
        "stores_camera_mapping": [],
        "raw_xml_samples": [],
        "parsed_data": [],
        "diagnosis": []
    }
    
    # Get active VMS servers
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    result["vms_servers"] = [{"id": v["id"], "name": v["name"], "base_url": v.get("base_url", "")} for v in vms_servers]
    
    if not vms_servers:
        result["diagnosis"].append("❌ HİÇ AKTİF VMS SUNUCUSU YOK!")
        return result
    
    # Get stores and their analytics camera mappings
    stores = await db.stores.find({}, {"_id": 0}).to_list(500)
    for store in stores:
        analytics_camera_id = store.get("analytics_camera_id")
        analytics_camera_ids = store.get("analytics_camera_ids", [])
        result["stores_camera_mapping"].append({
            "store_id": store["id"],
            "store_name": store["name"],
            "analytics_camera_id": analytics_camera_id,
            "analytics_camera_ids": analytics_camera_ids,
            "has_analytics_camera": bool(analytics_camera_id or analytics_camera_ids)
        })
    
    stores_with_analytics = [s for s in result["stores_camera_mapping"] if s["has_analytics_camera"]]
    if not stores_with_analytics:
        result["diagnosis"].append("❌ HİÇBİR MAĞAZAYA ANALİTİK KAMERA ATANMAMIŞ!")
        result["diagnosis"].append("📝 Mağaza düzenleme sayfasından 'Analytics Kamera' alanını doldurun.")
        return result
    else:
        result["diagnosis"].append(f"✅ {len(stores_with_analytics)} mağazada analytics kamera tanımlı")
    
    # Try to fetch raw data from each VMS
    for vms in vms_servers:
        try:
            # Face recognition search endpoint
            xml_data = await fetch_vms_data(vms, "/rsapi/modules/fr/searchevents?lastMinutes=30")
            
            if xml_data:
                # Save first 2000 chars of raw XML
                result["raw_xml_samples"].append({
                    "vms_name": vms["name"],
                    "endpoint": "/rsapi/modules/fr/searchevents?lastMinutes=30",
                    "xml_length": len(xml_data),
                    "xml_preview": xml_data[:2000] if len(xml_data) > 2000 else xml_data
                })
                
                # Parse it
                parsed = parse_analytics_xml(xml_data)
                cameras_data = parsed.get('cameras', [])
                
                result["parsed_data"].append({
                    "vms_name": vms["name"],
                    "cameras_count": len(cameras_data),
                    "cameras": cameras_data[:5]  # First 5 cameras
                })
                
                if cameras_data:
                    total_detections = sum(len(c.get('detections', [])) for c in cameras_data)
                    result["diagnosis"].append(f"✅ VMS '{vms['name']}': {len(cameras_data)} kamera, {total_detections} tespit")
                else:
                    result["diagnosis"].append(f"⚠️ VMS '{vms['name']}': XML geldi ama kamera verisi parse edilemedi")
            else:
                result["diagnosis"].append(f"❌ VMS '{vms['name']}': Veri çekilemedi (XML boş)")
                
        except Exception as e:
            result["diagnosis"].append(f"❌ VMS '{vms['name']}' HATA: {str(e)}")
    
    # Check if camera IDs match
    all_vms_camera_ids = set()
    for pd in result["parsed_data"]:
        for cam in pd.get("cameras", []):
            all_vms_camera_ids.add(cam.get("camera_id"))
    
    store_camera_ids = set()
    for store in stores_with_analytics:
        if store.get("analytics_camera_id"):
            store_camera_ids.add(store["analytics_camera_id"])
        for cid in store.get("analytics_camera_ids", []):
            store_camera_ids.add(cid)
    
    matching_cameras = all_vms_camera_ids.intersection(store_camera_ids)
    if matching_cameras:
        result["diagnosis"].append(f"✅ {len(matching_cameras)} kamera ID'si eşleşiyor: {list(matching_cameras)[:5]}")
    else:
        result["diagnosis"].append(f"❌ KAMERA ID EŞLEŞMESİ YOK!")
        result["diagnosis"].append(f"   VMS'deki kamera ID'leri: {list(all_vms_camera_ids)[:5]}")


@api_router.get("/debug/trigger-analytics-snapshot")
async def trigger_analytics_snapshot():
    """Manually trigger analytics snapshot collection - FOR TESTING"""
    from vms_utils import fetch_vms_data, parse_analytics_xml
    
    result = {
        "step1_vms_fetch": {},
        "step2_parse": {},
        "step3_camera_matching": {},
        "step4_snapshot_save": {}
    }
    
    try:
        # Step 1: Fetch from VMS
        vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
        stores = await db.stores.find({}, {"_id": 0}).to_list(500)
        
        vms_data = {}
        for vms in vms_servers:
            xml_data = await fetch_vms_data(vms, "/rsapi/modules/fr/searchevents?lastMinutes=5")
            result["step1_vms_fetch"][vms["name"]] = {
                "xml_length": len(xml_data) if xml_data else 0,
                "xml_preview": xml_data[:500] if xml_data else "NO DATA"
            }
            
            if xml_data:
                # Step 2: Parse
                parsed = parse_analytics_xml(xml_data)
                cameras_list = parsed.get('cameras', [])
                result["step2_parse"][vms["name"]] = {
                    "cameras_count": len(cameras_list),
                    "cameras": [{
                        "camera_id": c.get("camera_id"),
                        "detections_count": len(c.get("detections", []))
                    } for c in cameras_list]
                }
                
                # Build vms_data dict
                for p in cameras_list:
                    events = []
                    for det in p.get('detections', []):
                        events.append({
                            'gender': det.get('gender', 'Unknown'),
                            'age': int(det.get('age', 0)) if det.get('age') else 0
                        })
                    vms_data[p["camera_id"]] = {
                        "camera_id": p["camera_id"],
                        "events": events
                    }
        
        result["step2_parse"]["vms_data_keys"] = list(vms_data.keys())
        
        # Step 3: Match with stores
        for store in stores:
            analytics_camera_ids = store.get("analytics_camera_ids", [])
            if store.get("analytics_camera_id") and store["analytics_camera_id"] not in analytics_camera_ids:
                analytics_camera_ids.append(store["analytics_camera_id"])
            
            matched_events = 0
            for cam_id in analytics_camera_ids:
                cam_data = vms_data.get(cam_id)
                if cam_data:
                    matched_events += len(cam_data.get("events", []))
            
            result["step3_camera_matching"][store["name"]] = {
                "store_analytics_camera_ids": analytics_camera_ids,
                "matched_events": matched_events
            }
        
        # Step 4: Actually trigger the collection
        from data_collector import collect_analytics_snapshot
        
        before_count = await db.analytics_snapshots.count_documents({})
        before_with_events = await db.analytics_snapshots.count_documents({"total_events": {"$gt": 0}})
        
        await collect_analytics_snapshot()
        
        after_count = await db.analytics_snapshots.count_documents({})
        after_with_events = await db.analytics_snapshots.count_documents({"total_events": {"$gt": 0}})
        
        latest = await db.analytics_snapshots.find(
            {}, 
            {"_id": 0, "store_name": 1, "total_events": 1, "timestamp": 1}
        ).sort("timestamp", -1).limit(4).to_list(4)
        
        result["step4_snapshot_save"] = {
            "before": {"total": before_count, "with_events": before_with_events},
            "after": {"total": after_count, "with_events": after_with_events},
            "new_created": after_count - before_count,
            "latest": latest
        }
        
        result["success"] = True
        return result
        
    except Exception as e:
        result["error"] = str(e)
        import traceback
        result["traceback"] = traceback.format_exc()
        return result


@api_router.get("/debug/cleanup-empty-analytics")
async def cleanup_empty_analytics():
    """Delete analytics snapshots with total_events = 0"""
    try:
        # Count before
        total_before = await db.analytics_snapshots.count_documents({})
        empty_before = await db.analytics_snapshots.count_documents({"total_events": 0})
        with_events = await db.analytics_snapshots.count_documents({"total_events": {"$gt": 0}})
        
        # Delete empty ones
        result = await db.analytics_snapshots.delete_many({"total_events": 0})
        
        # Count after
        total_after = await db.analytics_snapshots.count_documents({})
        
        return {
            "success": True,
            "before": {
                "total": total_before,
                "empty": empty_before,
                "with_events": with_events
            },
            "deleted": result.deleted_count,
            "after": {
                "total": total_after
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@api_router.get("/debug/cleanup-deleted-stores")
async def cleanup_deleted_stores():
    """Delete snapshots for stores that no longer exist"""
    try:
        # Get active store IDs
        active_stores = await db.stores.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
        active_store_ids = [s["id"] for s in active_stores]
        
        # Find snapshots with non-existent stores
        # Counter snapshots
        counter_deleted = await db.counter_snapshots.delete_many({"store_id": {"$nin": active_store_ids}})
        
        # Queue snapshots
        queue_deleted = await db.queue_snapshots.delete_many({"store_id": {"$nin": active_store_ids}})
        
        # Analytics snapshots
        analytics_deleted = await db.analytics_snapshots.delete_many({"store_id": {"$nin": active_store_ids}})
        
        # Daily summaries
        daily_deleted = await db.daily_summaries.delete_many({"store_id": {"$nin": active_store_ids}})
        
        return {
            "success": True,
            "active_stores": [{"id": s["id"], "name": s["name"]} for s in active_stores],
            "deleted": {
                "counter_snapshots": counter_deleted.deleted_count,
                "queue_snapshots": queue_deleted.deleted_count,
                "analytics_snapshots": analytics_deleted.deleted_count,
                "daily_summaries": daily_deleted.deleted_count
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@api_router.get("/debug/scheduled-reports-check")
async def debug_scheduled_reports_check():
    """Debug endpoint to check scheduled reports saved values"""
    reports = await db.scheduled_reports.find({}, {"_id": 0}).to_list(100)
    return {
        "total_reports": len(reports),
        "reports": [
            {
                "id": r.get("id"),
                "name": r.get("name"),
                "report_type": r.get("report_type"),
                "date_range": r.get("date_range"),
                "store_ids": r.get("store_ids"),
                "frequency": r.get("frequency")
            }
            for r in reports
        ]
    }


@api_router.get("/debug/test-report-generation")
async def debug_test_report_generation(
    date_range: str = "1d",
    report_type: str = "counter"
):
    """Debug endpoint to test report generation with specific date_range"""
    logger.info(f"DEBUG TEST: Generating {report_type} report with date_range={date_range}")
    
    filters = {
        "store_ids": [],
        "date_range": date_range,
        "hour_from": None,
        "hour_to": None,
        "gender_filter": None,
        "min_queue_length": 0
    }
    
    report_data = await generate_report_data(report_type, filters)
    
    return {
        "requested_date_range": date_range,
        "report_type": report_type,
        "data_count": len(report_data.get("data", [])),
        "report_type_label": report_data.get("type"),
        "sample_data": report_data.get("data", [])[:5],  # First 5 rows
        "total_in_sum": sum(item.get("Giriş", 0) for item in report_data.get("data", []) if item.get("Mağaza") != "TOPLAM") if report_type == "counter" else None
    }


@api_router.post("/debug/fix-null-date-ranges")
async def debug_fix_null_date_ranges(default_date_range: str = "1d"):
    """Fix scheduled reports with null date_range values"""
    # Find all reports with null date_range
    null_reports = await db.scheduled_reports.find(
        {"$or": [{"date_range": None}, {"date_range": {"$exists": False}}]},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    
    if not null_reports:
        return {"message": "No reports with null date_range found", "fixed_count": 0}
    
    # Update all null date_range to default
    result = await db.scheduled_reports.update_many(
        {"$or": [{"date_range": None}, {"date_range": {"$exists": False}}]},
        {"$set": {"date_range": default_date_range}}
    )
    
    return {
        "message": f"Fixed {result.modified_count} reports",
        "fixed_count": result.modified_count,
        "reports_fixed": [r["name"] for r in null_reports],
        "new_date_range": default_date_range
    }


@api_router.get("/debug/raw-counter-data")
async def debug_raw_counter_data():
    """Check raw counter_snapshots data in database"""
    from datetime import timedelta
    
    now = datetime.now(timezone.utc)
    
    # Get last 7 days raw data
    result = {}
    for i in range(7):
        date = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        
        # Count documents for this date
        count = await db.counter_snapshots.count_documents({"date": date})
        
        # Get sample data
        samples = await db.counter_snapshots.find(
            {"date": date}, 
            {"_id": 0, "store_id": 1, "store_name": 1, "total_in": 1, "total_out": 1, "hour": 1}
        ).sort("hour", -1).limit(3).to_list(3)
        
        # Get max total_in for this date
        pipeline = [
            {"$match": {"date": date}},
            {"$group": {
                "_id": "$store_id",
                "max_in": {"$max": "$total_in"},
                "max_out": {"$max": "$total_out"},
                "count": {"$sum": 1}
            }}
        ]
        aggregated = await db.counter_snapshots.aggregate(pipeline).to_list(10)
        
        result[date] = {
            "document_count": count,
            "samples": samples,
            "aggregated_by_store": aggregated
        }
    
    return result


@api_router.get("/debug/date-range-compare")
async def debug_date_range_compare():
    from datetime import timedelta
    
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    week_ago = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    
    result = {
        "current_time": now.isoformat(),
        "today": today,
        "week_ago": week_ago,
        "counter_snapshots": {},
        "analytics_snapshots": {},
        "queue_snapshots": {},
        "report_comparison": {}
    }
    
    # Check counter_snapshots by date
    dates_with_data = []
    for i in range(8):
        day_date = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        
        pipeline = [
            {"$match": {"date": day_date}},
            {"$sort": {"hour": -1, "minute": -1}},
            {"$group": {
                "_id": "$store_id",
                "total_in": {"$first": "$total_in"},
                "total_out": {"$first": "$total_out"}
            }}
        ]
        day_results = await db.counter_snapshots.aggregate(pipeline).to_list(100)
        
        day_total_in = sum(r.get("total_in", 0) for r in day_results)
        dates_with_data.append({
            "date": day_date,
            "store_count": len(day_results),
            "total_in": day_total_in,
            "total_out": sum(r.get("total_out", 0) for r in day_results)
        })
    
    result["counter_snapshots"]["by_date"] = dates_with_data
    result["counter_snapshots"]["week_total_in"] = sum(d["total_in"] for d in dates_with_data)
    
    # Check analytics_snapshots by date
    analytics_dates = []
    for i in range(8):
        day_date = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        
        pipeline = [
            {"$match": {"date": day_date}},
            {"$group": {
                "_id": "$store_id",
                "total_events": {"$sum": "$total_events"}
            }}
        ]
        day_results = await db.analytics_snapshots.aggregate(pipeline).to_list(100)
        
        day_total = sum(r.get("total_events", 0) for r in day_results)
        analytics_dates.append({
            "date": day_date,
            "store_count": len(day_results),
            "total_events": day_total
        })
    
    result["analytics_snapshots"]["by_date"] = analytics_dates
    result["analytics_snapshots"]["week_total"] = sum(d["total_events"] for d in analytics_dates)
    
    # Check queue_snapshots by date
    queue_dates = []
    for i in range(8):
        day_date = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        
        pipeline = [
            {"$match": {"date": day_date}},
            {"$group": {
                "_id": "$store_id",
                "total_queue": {"$sum": "$total_queue_length"}
            }}
        ]
        day_results = await db.queue_snapshots.aggregate(pipeline).to_list(100)
        
        day_total = sum(r.get("total_queue", 0) for r in day_results)
        queue_dates.append({
            "date": day_date,
            "store_count": len(day_results),
            "total_queue": day_total
        })
    
    result["queue_snapshots"]["by_date"] = queue_dates
    result["queue_snapshots"]["week_total"] = sum(d["total_queue"] for d in queue_dates)
    
    # Get 1 day and 1 week reports for counter
    try:
        report_1d = await get_counter_report(date_range="1d")
        result["report_comparison"]["counter_1d"] = {
            "total_in": report_1d.get("summary", {}).get("total_in", 0),
            "store_count": len(report_1d.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["counter_1d"] = {"error": str(e)}
    
    try:
        report_1w = await get_counter_report(date_range="1w")
        result["report_comparison"]["counter_1w"] = {
            "total_in": report_1w.get("summary", {}).get("total_in", 0),
            "store_count": len(report_1w.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["counter_1w"] = {"error": str(e)}
    
    # Get analytics reports
    try:
        analytics_1d = await get_analytics_report(date_range="1d")
        result["report_comparison"]["analytics_1d"] = {
            "total_detections": analytics_1d.get("summary", {}).get("total_detections", 0),
            "store_count": len(analytics_1d.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["analytics_1d"] = {"error": str(e)}
    
    try:
        analytics_1w = await get_analytics_report(date_range="1w")
        result["report_comparison"]["analytics_1w"] = {
            "total_detections": analytics_1w.get("summary", {}).get("total_detections", 0),
            "store_count": len(analytics_1w.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["analytics_1w"] = {"error": str(e)}
    
    # Get queue reports
    try:
        queue_1d = await get_queue_report(date_range="1d")
        result["report_comparison"]["queue_1d"] = {
            "total_queue": queue_1d.get("summary", {}).get("total_queue_length", 0),
            "store_count": len(queue_1d.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["queue_1d"] = {"error": str(e)}
    
    try:
        queue_1w = await get_queue_report(date_range="1w")
        result["report_comparison"]["queue_1w"] = {
            "total_queue": queue_1w.get("summary", {}).get("total_queue_length", 0),
            "store_count": len(queue_1w.get("stores", []))
        }
    except Exception as e:
        result["report_comparison"]["queue_1w"] = {"error": str(e)}
    
    return result


@api_router.get("/sentry-test")
async def test_sentry(user: dict = Depends(require_admin)):
    """Test Sentry error tracking (admin only)"""
    import sentry_sdk
    
    if not os.environ.get("SENTRY_DSN"):
        return {"status": "error", "message": "SENTRY_DSN not configured"}
    
    try:
        # Capture a test message
        sentry_sdk.capture_message("VMS360 Sentry test - this is a test alert", level="info")
        return {
            "status": "success", 
            "message": "Test mesajı Sentry'ye gönderildi. Dashboard'u kontrol edin."
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@api_router.get("/cache/stats")
async def get_cache_stats(user: dict = Depends(require_admin)):
    """Get Redis cache statistics (admin only)"""
    from cache import get_cache_stats
    return await get_cache_stats()


@api_router.post("/cache/clear")
async def clear_cache(user: dict = Depends(require_admin)):
    """Clear all Redis cache (admin only)"""
    from cache import delete_cached
    await delete_cached("*")
    return {"message": "Cache temizlendi"}


@api_router.get("/alerts/settings")
async def get_alert_settings(user: dict = Depends(require_admin)):
    """Get alert settings (admin only)"""
    settings = await db.alert_settings.find_one({}, {"_id": 0})
    if not settings:
        settings = {
            "health_alert_emails": [],
            "offline_threshold_minutes": 30,
            "enabled": True
        }
    return settings


@api_router.put("/alerts/settings")
async def update_alert_settings(
    settings: dict,
    user: dict = Depends(require_admin)
):
    """Update alert settings (admin only)"""
    update_data = {
        "health_alert_emails": settings.get("health_alert_emails", []),
        "offline_threshold_minutes": settings.get("offline_threshold_minutes", 30),
        "enabled": settings.get("enabled", True),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.alert_settings.update_one(
        {},
        {"$set": update_data},
        upsert=True
    )
    return {"message": "Ayarlar güncellendi", "settings": update_data}


@api_router.post("/alerts/test")
async def test_alert_email(user: dict = Depends(require_admin)):
    """Send a test alert email (admin only)"""
    from data_collector import send_health_alert_email
    
    test_offline = [{"store_id": "test", "store_name": "Test Mağaza", "last_data": "Test"}]
    test_online = []
    
    await send_health_alert_email(test_offline, test_online)
    return {"message": "Test email gönderildi"}


@api_router.post("/vms", response_model=VMSServer)
async def create_vms(input: VMSServerCreate):
    vms_obj = VMSServer(**input.model_dump())
    doc = vms_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.vms_servers.insert_one(doc)
    return vms_obj

@api_router.get("/vms", response_model=List[VMSServer])
async def get_vms_list():
    servers = await db.vms_servers.find({}, {"_id": 0}).to_list(100)
    for s in servers:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
    return servers

@api_router.get("/vms/{vms_id}", response_model=VMSServer)
async def get_vms(vms_id: str):
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    if isinstance(server.get('created_at'), str):
        server['created_at'] = datetime.fromisoformat(server['created_at'])
    return server

@api_router.put("/vms/{vms_id}", response_model=VMSServer)
async def update_vms(vms_id: str, input: VMSServerUpdate):
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    result = await db.vms_servers.update_one({"id": vms_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="VMS not found")
    return await get_vms(vms_id)

@api_router.delete("/vms/{vms_id}")
async def delete_vms(vms_id: str):
    result = await db.vms_servers.delete_one({"id": vms_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="VMS not found")
    return {"status": "deleted"}

@api_router.get("/vms/{vms_id}/test")
async def test_vms_connection(vms_id: str):
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    data = await fetch_vms_data(server, "/rsapi/modules/counter/getstats")
    if data:
        return {"status": "connected", "message": "VMS bağlantısı başarılı"}
    return {"status": "error", "message": "VMS bağlantısı kurulamadı"}

@api_router.get("/vms/{vms_id}/cameras")
async def fetch_vms_cameras(vms_id: str):
    """Fetch available cameras from VMS server with proper names from /rsapi/cameras"""
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    # First, get ALL cameras with names from /rsapi/cameras endpoint
    camera_list_xml = await fetch_vms_data(server, "/rsapi/cameras")
    all_cameras = {}
    
    if camera_list_xml:
        parsed_list = parse_camera_list_xml(camera_list_xml)
        for cam in parsed_list.get('cameras', []):
            all_cameras[cam['camera_id']] = {
                'camera_id': cam['camera_id'],
                'name': cam['name'],
                'description': cam.get('description', ''),
                'disabled': cam.get('disabled', False),
                'model': cam.get('model', ''),
                'has_counter': False,
                'has_queue': False,
                'has_analytics': False,
                'type': 'general'
            }
    
    # Check which cameras are used in counter module
    counter_data = await fetch_vms_data(server, "/rsapi/modules/counter/getstats")
    if counter_data:
        parsed = parse_counter_xml(counter_data)
        for cam in parsed.get('cameras', []):
            cam_id = cam.get('camera_id')
            if cam_id:
                # Get in/out counts from counters array
                counters = cam.get('counters', [{}])
                in_count = counters[0].get('in_count', 0) if counters else 0
                out_count = counters[0].get('out_count', 0) if counters else 0
                
                if cam_id in all_cameras:
                    all_cameras[cam_id]['has_counter'] = True
                    all_cameras[cam_id]['type'] = 'counter'
                    all_cameras[cam_id]['in_count'] = in_count
                    all_cameras[cam_id]['out_count'] = out_count
                    all_cameras[cam_id]['last_reset'] = cam.get('last_reset', '')
                else:
                    all_cameras[cam_id] = {
                        'camera_id': cam_id,
                        'name': cam.get('camera_name') or f"Sayaç Kamera {cam_id[:8]}",
                        'has_counter': True,
                        'has_queue': False,
                        'has_analytics': False,
                        'type': 'counter',
                        'in_count': in_count,
                        'out_count': out_count
                    }
    
    # Check which cameras are used in queue module
    queue_data = await fetch_vms_data(server, "/rsapi/modules/queue/getstats")
    if queue_data:
        parsed = parse_queue_xml(queue_data)
        for cam in parsed.get('cameras', []):
            cam_id = cam.get('camera_id')
            if cam_id:
                if cam_id in all_cameras:
                    all_cameras[cam_id]['has_queue'] = True
                    all_cameras[cam_id]['zones'] = cam.get('zones', [])
                    if all_cameras[cam_id]['type'] == 'general':
                        all_cameras[cam_id]['type'] = 'queue'
                else:
                    all_cameras[cam_id] = {
                        'camera_id': cam_id,
                        'name': cam.get('camera_name') or f"Kuyruk Kamera {cam_id[:8]}",
                        'has_counter': False,
                        'has_queue': True,
                        'has_analytics': False,
                        'type': 'queue',
                        'zones': cam.get('zones', [])
                    }
    
    # Check which cameras are used in analytics/FR module
    analytics_data = await fetch_vms_data(server, "/rsapi/modules/fr/analytics/getstats")
    if analytics_data:
        parsed = parse_analytics_xml(analytics_data)
        for cam in parsed.get('cameras', []):
            cam_id = cam.get('camera_id')
            if cam_id:
                if cam_id in all_cameras:
                    all_cameras[cam_id]['has_analytics'] = True
                    if all_cameras[cam_id]['type'] == 'general':
                        all_cameras[cam_id]['type'] = 'analytics'
                else:
                    all_cameras[cam_id] = {
                        'camera_id': cam_id,
                        'name': cam.get('camera_name') or f"Analitik Kamera {cam_id[:8]}",
                        'has_counter': False,
                        'has_queue': False,
                        'has_analytics': True,
                        'type': 'analytics'
                    }
    
    # Convert to list and sort
    cameras = list(all_cameras.values())
    cameras.sort(key=lambda c: (c.get('disabled', False), c.get('name', '')))
    
    return {
        "vms_id": vms_id,
        "vms_name": server.get("name", ""),
        "cameras": cameras,
        "total": len(cameras)
    }

class ImportCamerasRequest(BaseModel):
    store_id: str
    cameras: List[dict]

@api_router.post("/vms/{vms_id}/import-cameras")
async def import_vms_cameras(vms_id: str, request: ImportCamerasRequest):
    """Import cameras from VMS and save to database"""
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    store = await db.stores.find_one({"id": request.store_id}, {"_id": 0})
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    imported = 0
    skipped = 0
    
    for cam_data in request.cameras:
        # Check if camera already exists
        existing = await db.cameras.find_one({
            "camera_vms_id": cam_data["camera_id"],
            "store_id": request.store_id
        })
        
        if existing:
            skipped += 1
            continue
        
        # Create new camera
        camera = Camera(
            store_id=request.store_id,
            camera_vms_id=cam_data["camera_id"],
            name=cam_data.get("name", f"Kamera {cam_data['camera_id'][:8]}"),
            type=cam_data.get("type", "counter")
        )
        doc = camera.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.cameras.insert_one(doc)
        imported += 1
    
    return {
        "status": "success",
        "imported": imported,
        "skipped": skipped,
        "message": f"{imported} kamera eklendi, {skipped} kamera zaten mevcut"
    }

@api_router.post("/vms/{vms_id}/sync-cameras")
async def sync_vms_cameras(vms_id: str, store_id: Optional[str] = None):
    """Fetch and automatically import all cameras from VMS"""
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    # Get cameras from VMS
    cameras_response = await fetch_vms_cameras(vms_id)
    vms_cameras = cameras_response["cameras"]
    
    if not vms_cameras:
        return {"status": "warning", "message": "VMS'de kamera bulunamadı", "imported": 0}
    
    # If store_id provided, use that store; otherwise find stores linked to this VMS
    if store_id:
        stores = [await db.stores.find_one({"id": store_id}, {"_id": 0})]
        stores = [s for s in stores if s]
    else:
        stores = await db.stores.find({"vms_id": vms_id}, {"_id": 0}).to_list(100)
    
    if not stores:
        return {"status": "warning", "message": "Bu VMS'e bağlı mağaza bulunamadı", "imported": 0}
    
    total_imported = 0
    total_skipped = 0
    
    for store in stores:
        for cam_data in vms_cameras:
            # Check if camera already exists for this store
            existing = await db.cameras.find_one({
                "camera_vms_id": cam_data["camera_id"],
                "store_id": store["id"]
            })
            
            if existing:
                total_skipped += 1
                continue
            
            # Create camera with better naming
            cam_type = cam_data.get("type", "counter")
            type_names = {"counter": "Sayaç", "queue": "Kuyruk", "analytics": "Analitik"}
            cam_name = f"{type_names.get(cam_type, 'Kamera')} - {cam_data['camera_id'][:8]}"
            
            camera = Camera(
                store_id=store["id"],
                camera_vms_id=cam_data["camera_id"],
                name=cam_name,
                type=cam_type
            )
            doc = camera.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.cameras.insert_one(doc)
            total_imported += 1
    
    return {
        "status": "success",
        "imported": total_imported,
        "skipped": total_skipped,
        "stores_count": len(stores),
        "message": f"{total_imported} kamera eklendi ({len(stores)} mağazaya)"
    }


@api_router.post("/vms/{vms_id}/update-camera-names")
async def update_camera_names_from_vms(vms_id: str):
    """Update camera names in database from VMS"""
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    # Get cameras with real names from VMS
    camera_list_xml = await fetch_vms_data(server, "/rsapi/cameras")
    if not camera_list_xml:
        return {"status": "error", "message": "VMS'den kamera listesi alınamadı", "updated": 0}
    
    parsed_list = parse_camera_list_xml(camera_list_xml)
    vms_cameras = {cam['camera_id']: cam for cam in parsed_list.get('cameras', [])}
    
    if not vms_cameras:
        return {"status": "error", "message": "VMS'de kamera bulunamadı", "updated": 0}
    
    # Update cameras in database
    updated = 0
    for cam_id, cam_data in vms_cameras.items():
        result = await db.cameras.update_many(
            {"camera_vms_id": cam_id},
            {"$set": {"name": cam_data['name']}}
        )
        updated += result.modified_count
    
    return {
        "status": "success",
        "updated": updated,
        "vms_cameras_found": len(vms_cameras),
        "message": f"{updated} kamera ismi güncellendi"
    }


@api_router.post("/vms/{vms_id}/sync-all-cameras")
async def sync_all_cameras_from_vms(vms_id: str):
    """
    Full camera synchronization from VMS:
    1. Fetch all cameras from VMS
    2. Add new cameras to database
    3. Update existing camera names
    4. Mark cameras not in VMS as inactive
    5. Get module info (counter/queue/analytics capabilities)
    """
    server = await db.vms_servers.find_one({"id": vms_id}, {"_id": 0})
    if not server:
        raise HTTPException(status_code=404, detail="VMS not found")
    
    # Get cameras with real names from VMS
    camera_list_xml = await fetch_vms_data(server, "/rsapi/cameras")
    if not camera_list_xml:
        return {"status": "error", "message": "VMS'den kamera listesi alınamadı"}
    
    parsed_list = parse_camera_list_xml(camera_list_xml)
    vms_cameras = {cam['camera_id']: cam for cam in parsed_list.get('cameras', [])}
    
    if not vms_cameras:
        return {"status": "warning", "message": "VMS'de kamera bulunamadı"}
    
    # Get module capabilities
    counter_cameras = set()
    queue_cameras = set()
    analytics_cameras = set()
    
    # Check counter module
    counter_data = await fetch_vms_data(server, "/rsapi/modules/counter/getstats")
    if counter_data:
        parsed = parse_counter_xml(counter_data)
        for cam in parsed:
            counter_cameras.add(cam.get('camera_id'))
    
    # Check queue module
    queue_data = await fetch_vms_data(server, "/rsapi/modules/queue/getstats")
    if queue_data:
        parsed = parse_queue_xml(queue_data)
        for cam in parsed:
            queue_cameras.add(cam.get('camera_id'))
    
    # Check analytics module
    analytics_data = await fetch_vms_data(server, "/rsapi/modules/fr/analytics/getstats")
    if analytics_data:
        parsed = parse_counter_xml(analytics_data)
        for cam in parsed:
            analytics_cameras.add(cam.get('camera_id'))
    
    # Get existing cameras in database
    existing_cameras = await db.cameras.find({"vms_id": vms_id}, {"_id": 0}).to_list(500)
    existing_by_vms_id = {c["camera_vms_id"]: c for c in existing_cameras}
    
    added = 0
    updated = 0
    deactivated = 0
    
    # Process VMS cameras
    for cam_id, cam_data in vms_cameras.items():
        has_counter = cam_id in counter_cameras
        has_queue = cam_id in queue_cameras
        has_analytics = cam_id in analytics_cameras
        
        # Determine camera type
        if has_counter:
            cam_type = "counter"
        elif has_queue:
            cam_type = "queue"
        elif has_analytics:
            cam_type = "analytics"
        else:
            cam_type = "general"
        
        camera_doc = {
            "name": cam_data['name'],
            "description": cam_data.get('description', ''),
            "disabled": cam_data.get('disabled', False),
            "model": cam_data.get('model', ''),
            "has_counter": has_counter,
            "has_queue": has_queue,
            "has_analytics": has_analytics,
            "type": cam_type,
            "vms_id": vms_id,
            "is_active": not cam_data.get('disabled', False),
            "last_sync": datetime.now(timezone.utc).isoformat()
        }
        
        if cam_id in existing_by_vms_id:
            # Update existing camera
            await db.cameras.update_one(
                {"camera_vms_id": cam_id},
                {"$set": camera_doc}
            )
            updated += 1
        else:
            # Add new camera
            import uuid
            camera_doc["id"] = str(uuid.uuid4())
            camera_doc["camera_vms_id"] = cam_id
            camera_doc["store_id"] = ""  # Not assigned to any store yet
            camera_doc["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.cameras.insert_one(camera_doc)
            added += 1
    
    # Mark cameras not in VMS as inactive
    for cam_vms_id, existing in existing_by_vms_id.items():
        if cam_vms_id not in vms_cameras:
            await db.cameras.update_one(
                {"camera_vms_id": cam_vms_id},
                {"$set": {"is_active": False, "last_sync": datetime.now(timezone.utc).isoformat()}}
            )
            deactivated += 1
    
    return {
        "status": "success",
        "vms_cameras_found": len(vms_cameras),
        "added": added,
        "updated": updated,
        "deactivated": deactivated,
        "counter_cameras": len(counter_cameras),
        "queue_cameras": len(queue_cameras),
        "analytics_cameras": len(analytics_cameras),
        "message": f"Senkronizasyon tamamlandı: {added} eklendi, {updated} güncellendi, {deactivated} devre dışı"
    }


@api_router.post("/vms/sync-all")
async def sync_all_vms_cameras():
    """Sync cameras from all active VMS servers"""
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    if not vms_servers:
        return {"status": "warning", "message": "Aktif VMS sunucusu bulunamadı"}
    
    total_results = {
        "vms_count": len(vms_servers),
        "total_added": 0,
        "total_updated": 0,
        "total_deactivated": 0,
        "details": []
    }
    
    for vms in vms_servers:
        try:
            result = await sync_all_cameras_from_vms(vms["id"])
            total_results["total_added"] += result.get("added", 0)
            total_results["total_updated"] += result.get("updated", 0)
            total_results["total_deactivated"] += result.get("deactivated", 0)
            total_results["details"].append({
                "vms_id": vms["id"],
                "vms_name": vms["name"],
                "status": result.get("status"),
                "message": result.get("message")
            })
        except Exception as e:
            total_results["details"].append({
                "vms_id": vms["id"],
                "vms_name": vms["name"],
                "status": "error",
                "message": str(e)
            })
    
    total_results["message"] = f"Tüm VMS'ler senkronize edildi: {total_results['total_added']} eklendi, {total_results['total_updated']} güncellendi"
    return total_results


# ============== LOCATION ENDPOINTS (Now handled by routers/locations.py) ==============
# Location endpoints (regions, cities, districts, hierarchy) moved to routers/locations.py

# ============== STORE ENDPOINTS (Now handled by routers/stores.py) ==============
# Store endpoints moved to routers/stores.py

# ============== CAMERA ENDPOINTS (Now handled by routers/cameras.py) ==============
# Camera endpoints moved to routers/cameras.py

# ============== SETTINGS ENDPOINTS (Now handled by routers/settings.py) ==============
# Settings endpoints moved to routers/settings.py

# ============== LIVE DATA HELPERS (for internal use) ==============

async def _fetch_live_counter_data(store_ids: Optional[str] = None, allowed_stores: Optional[set] = None):
    """Internal helper to fetch live counter data without auth dependency"""
    result = []
    
    # Get stores with permission filtering
    store_query = {}
    if store_ids:
        requested_ids = store_ids.split(",") if isinstance(store_ids, str) else store_ids
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        store_query["id"] = {"$in": requested_ids}
    elif allowed_stores is not None:
        if not allowed_stores:
            return []
        store_query["id"] = {"$in": list(allowed_stores)}
    
    stores = await db.stores.find(store_query, {"_id": 0}).to_list(100)
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    vms_dict = {v["id"]: v for v in vms_servers}
    
    vms_data = {}
    for vms_id, vms in vms_dict.items():
        xml_data = await fetch_vms_data(vms, "/rsapi/modules/counter/getstats")
        if xml_data:
            parsed = parse_counter_xml(xml_data)
            for p in parsed.get('cameras', []):
                in_count = sum(c.get('in_count', 0) for c in p.get('counters', []))
                out_count = sum(c.get('out_count', 0) for c in p.get('counters', []))
                vms_data[p["camera_id"]] = {
                    "camera_id": p["camera_id"],
                    "in_count": in_count,
                    "out_count": out_count,
                    "last_reset": p.get("last_reset", "")
                }
    
    for store in stores:
        counter_camera_ids = store.get("counter_camera_ids", [])
        if store.get("counter_camera_id") and store["counter_camera_id"] not in counter_camera_ids:
            counter_camera_ids.append(store["counter_camera_id"])
        
        total_in = 0
        total_out = 0
        camera_details = []
        
        for cam_id in counter_camera_ids:
            cam_data = vms_data.get(cam_id)
            if cam_data:
                total_in += cam_data.get("in_count", 0)
                total_out += cam_data.get("out_count", 0)
                camera_details.append({
                    "camera_id": cam_id,
                    "in_count": cam_data.get("in_count", 0),
                    "out_count": cam_data.get("out_count", 0)
                })
        
        current_visitors = max(0, total_in - total_out)
        occupancy_percent = (current_visitors / store["capacity"] * 100) if store["capacity"] > 0 else 0
        
        district = await db.districts.find_one({"id": store.get("district_id")}, {"_id": 0})
        city = None
        region = None
        if district:
            city = await db.cities.find_one({"id": district.get("city_id")}, {"_id": 0})
            if city:
                region = await db.regions.find_one({"id": city.get("region_id")}, {"_id": 0})
        
        result.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "district_id": store.get("district_id"),
            "district_name": district["name"] if district else "",
            "city_id": city["id"] if city else "",
            "city_name": city["name"] if city else "",
            "region_id": region["id"] if region else "",
            "region_name": region["name"] if region else "",
            "total_in": total_in,
            "total_out": total_out,
            "current_visitors": current_visitors,
            "capacity": store["capacity"],
            "occupancy_percent": round(occupancy_percent, 1),
            "status": "critical" if occupancy_percent >= 95 else "warning" if occupancy_percent >= 80 else "normal",
            "camera_details": camera_details
        })
    
    return result


async def _fetch_live_queue_data(store_ids: Optional[str] = None, allowed_stores: Optional[set] = None):
    """Internal helper to fetch live queue data without auth dependency"""
    result = []
    
    store_query = {}
    if store_ids:
        requested_ids = store_ids.split(",") if isinstance(store_ids, str) else store_ids
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        store_query["id"] = {"$in": requested_ids}
    elif allowed_stores is not None:
        if not allowed_stores:
            return []
        store_query["id"] = {"$in": list(allowed_stores)}
    
    stores = await db.stores.find(store_query, {"_id": 0}).to_list(100)
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    vms_dict = {v["id"]: v for v in vms_servers}
    
    vms_data = {}
    for vms_id, vms in vms_dict.items():
        xml_data = await fetch_vms_data(vms, "/rsapi/modules/queue/getstats")
        if xml_data:
            parsed = parse_queue_xml(xml_data)
            for p in parsed.get('cameras', []):
                vms_data[p["camera_id"]] = p
    
    camera_names = {}
    for vms_id, vms in vms_dict.items():
        camera_list_xml = await fetch_vms_data(vms, "/rsapi/cameras")
        if camera_list_xml:
            parsed_list = parse_camera_list_xml(camera_list_xml)
            for cam in parsed_list.get('cameras', []):
                camera_names[cam['camera_id']] = cam['name']
    
    for store in stores:
        queue_camera_ids = store.get("queue_camera_ids", [])
        if store.get("queue_camera_id") and store["queue_camera_id"] not in queue_camera_ids:
            queue_camera_ids.append(store["queue_camera_id"])
        
        zones = []
        total_queue = 0
        camera_details = []
        
        for cam_id in queue_camera_ids:
            cam_data = vms_data.get(cam_id)
            cam_name = camera_names.get(cam_id, f"Kamera {cam_id[:8]}")
            
            if cam_data:
                cam_zones = cam_data.get("zones", [])
                for zone in cam_zones:
                    total_queue += zone.get("queue_length", 0)
                    zones.append({
                        "camera_id": cam_id,
                        "camera_name": cam_name,
                        "zone_index": zone.get("zone_index", 0),
                        "queue_length": zone.get("queue_length", 0),
                        "is_queue": zone.get("is_queue", False)
                    })
                camera_details.append({
                    "camera_id": cam_id,
                    "camera_name": cam_name,
                    "zones": cam_zones
                })
        
        district = await db.districts.find_one({"id": store.get("district_id")}, {"_id": 0})
        city = None
        region = None
        if district:
            city = await db.cities.find_one({"id": district.get("city_id")}, {"_id": 0})
            if city:
                region = await db.regions.find_one({"id": city.get("region_id")}, {"_id": 0})
        
        result.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "district_id": store.get("district_id"),
            "district_name": district["name"] if district else "",
            "city_id": city["id"] if city else "",
            "city_name": city["name"] if city else "",
            "region_id": region["id"] if region else "",
            "region_name": region["name"] if region else "",
            "total_queue_length": total_queue,
            "zones": zones,
            "queue_threshold": store.get("queue_threshold", 5),
            "status": "critical" if total_queue >= store.get("queue_threshold", 5) * 2 else "warning" if total_queue >= store.get("queue_threshold", 5) else "normal",
            "camera_details": camera_details
        })
    
    return result


# ============== LIVE DATA ENDPOINTS ==============

@api_router.get("/live/counter")
async def get_live_counter_data(
    store_ids: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get live people counter data for all or specific stores"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    return await _fetch_live_counter_data(store_ids, allowed_stores)

@api_router.get("/live/queue")
async def get_live_queue_data(
    store_ids: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get live queue data for all or specific stores - supports multiple cameras"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    return await _fetch_live_queue_data(store_ids, allowed_stores)

# Helper function for analytics data (can be called internally)
async def _fetch_analytics_data(
    store_ids: Optional[str] = None,
    time_from: Optional[str] = None,
    time_to: Optional[str] = None,
    gender: Optional[str] = None,
    from_age: Optional[int] = None,
    to_age: Optional[int] = None,
    allowed_camera_ids: Optional[List[str]] = None,
    last_minutes: Optional[int] = None  # Added for date_range support
):
    """Internal helper to fetch analytics data from VMS"""
    result = {
        "total_events": 0,
        "gender_distribution": {"Male": 0, "Female": 0},
        "age_distribution": {
            "0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0
        },
        "events": []
    }
    
    # Get VMS servers
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    for vms in vms_servers:
        # Build query params
        params = []
        if time_from:
            params.append(f"timeFrom={time_from}")
        if time_to:
            params.append(f"timeTo={time_to}")
        if gender:
            params.append(f"gender={gender}")
        if from_age:
            params.append(f"fromAge={from_age}")
        if to_age:
            params.append(f"toAge={to_age}")
        
        # If no time params, use lastMinutes (default 1440 = 1 day)
        if not time_from and not time_to:
            mins = last_minutes if last_minutes else 1440  # Default 1 day instead of 60 minutes
            params.append(f"lastMinutes={mins}")
        
        query_string = "&".join(params)
        endpoint = f"/rsapi/modules/fr/searchevents?{query_string}"
        
        xml_data = await fetch_vms_data(vms, endpoint)
        if xml_data:
            parsed = parse_analytics_xml(xml_data)
            # parse_analytics_xml returns {'cameras': [...]} where each camera has 'detections'
            for camera in parsed.get('cameras', []):
                camera_id = camera.get('camera_id')
                
                # Filter by allowed camera IDs if specified
                if allowed_camera_ids is not None and camera_id not in allowed_camera_ids:
                    continue
                
                for event in camera.get('detections', []):
                    # Gender distribution - only count Male and Female
                    gender_val = event.get("gender", "")
                    if gender_val == "Male":
                        result["gender_distribution"]["Male"] += 1
                        result["total_events"] += 1
                    elif gender_val == "Female":
                        result["gender_distribution"]["Female"] += 1
                        result["total_events"] += 1
                    # Skip Unknown gender
                    
                    # Age distribution (only for valid genders)
                    if gender_val in ["Male", "Female"]:
                        age = event.get("age", 0)
                        if age < 18:
                            result["age_distribution"]["0-17"] += 1
                        elif age < 25:
                            result["age_distribution"]["18-24"] += 1
                        elif age < 35:
                            result["age_distribution"]["25-34"] += 1
                        elif age < 45:
                            result["age_distribution"]["35-44"] += 1
                        elif age < 55:
                            result["age_distribution"]["45-54"] += 1
                        else:
                            result["age_distribution"]["55+"] += 1
                    
                    event['camera_id'] = camera_id
                    result["events"].append(event)
    
    return result

@api_router.get("/live/analytics")
async def get_live_analytics_data(
    store_ids: Optional[str] = None,
    time_from: Optional[str] = None,
    time_to: Optional[str] = None,
    gender: Optional[str] = None,
    from_age: Optional[int] = None,
    to_age: Optional[int] = None,
    user: dict = Depends(require_auth)
):
    """Get analytics data (age/gender) for stores"""
    from permissions import get_user_allowed_stores
    
    # Get user's allowed stores
    allowed_stores = await get_user_allowed_stores(user)
    
    # If user has no access to any store, return empty
    if allowed_stores is not None and not allowed_stores:
        return {
            "total_events": 0,
            "gender_distribution": {"Male": 0, "Female": 0},
            "age_distribution": {
                "0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0
            },
            "events": []
        }
    
    # Get analytics camera IDs for the selected stores
    allowed_camera_ids = None
    if store_ids:
        requested_ids = store_ids.split(",")
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        
        # Get camera IDs from requested stores
        stores = await db.stores.find({"id": {"$in": requested_ids}}, {"_id": 0}).to_list(100)
        camera_ids = []
        for store in stores:
            # Get from both old (single) and new (multiple) fields
            analytics_ids = store.get("analytics_camera_ids", [])
            if store.get("analytics_camera_id") and store["analytics_camera_id"] not in analytics_ids:
                analytics_ids.append(store["analytics_camera_id"])
            camera_ids.extend(analytics_ids)
        
        if camera_ids:
            allowed_camera_ids = camera_ids
    
    return await _fetch_analytics_data(store_ids, time_from, time_to, gender, from_age, to_age, allowed_camera_ids)


@api_router.get("/live/analytics/stores")
async def get_live_analytics_by_store(
    store_ids: Optional[str] = None,
    time_from: Optional[str] = None,
    time_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get analytics data (age/gender) per store - supports multiple cameras"""
    from permissions import get_user_allowed_stores
    
    result = []
    
    # Get user's allowed stores
    allowed_stores = await get_user_allowed_stores(user)
    
    # Get stores with permission filtering
    store_query = {}
    if store_ids:
        requested_ids = store_ids.split(",")
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        store_query["id"] = {"$in": requested_ids}
    elif allowed_stores is not None:
        if not allowed_stores:
            return []  # No access
        store_query["id"] = {"$in": list(allowed_stores)}
    
    stores = await db.stores.find(store_query, {"_id": 0}).to_list(100)
    
    # Get VMS servers
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    vms_dict = {v["id"]: v for v in vms_servers}
    
    # Get camera names from VMS
    camera_names = {}
    for vms_id, vms in vms_dict.items():
        camera_list_xml = await fetch_vms_data(vms, "/rsapi/cameras")
        if camera_list_xml:
            parsed_list = parse_camera_list_xml(camera_list_xml)
            for cam in parsed_list.get('cameras', []):
                camera_names[cam['camera_id']] = cam['name']
    
    # Fetch analytics data from all VMS servers
    all_events = []
    for vms in vms_servers:
        endpoint = "/rsapi/modules/fr/searchevents?lastMinutes=1440"  # Default 1 day to match /live/analytics
        if time_from:
            endpoint = f"/rsapi/modules/fr/searchevents?timeFrom={time_from}"
            if time_to:
                endpoint += f"&timeTo={time_to}"
        
        xml_data = await fetch_vms_data(vms, endpoint)
        if xml_data:
            parsed = parse_analytics_xml(xml_data)
            # parse_analytics_xml returns {'cameras': [...]} where each camera has 'detections'
            for camera in parsed.get('cameras', []):
                camera_id = camera.get('camera_id')
                for detection in camera.get('detections', []):
                    detection['camera_id'] = camera_id
                    detection['vms_id'] = vms["id"]
                    all_events.append(detection)
    
    # Build result per store
    for store in stores:
        # Get location info
        district = await db.districts.find_one({"id": store.get("district_id")}, {"_id": 0})
        city = None
        region = None
        if district:
            city = await db.cities.find_one({"id": district.get("city_id")}, {"_id": 0})
            if city:
                region = await db.regions.find_one({"id": city.get("region_id")}, {"_id": 0})
        
        # Get camera IDs from both old (single) and new (multiple) fields
        analytics_camera_ids = store.get("analytics_camera_ids", [])
        # Backward compatibility: add old single camera if exists
        if store.get("analytics_camera_id") and store["analytics_camera_id"] not in analytics_camera_ids:
            analytics_camera_ids.append(store["analytics_camera_id"])
        
        # Filter events for this store's analytics cameras
        store_events = []
        camera_details = []
        
        if analytics_camera_ids:
            for cam_id in analytics_camera_ids:
                cam_events = [e for e in all_events if e.get("camera_id") == cam_id]
                store_events.extend(cam_events)
                
                # Calculate per-camera stats
                cam_male = sum(1 for e in cam_events if e.get("gender") == "Male")
                cam_female = sum(1 for e in cam_events if e.get("gender") == "Female")
                camera_details.append({
                    "camera_id": cam_id,
                    "camera_name": camera_names.get(cam_id, f"Kamera {cam_id[:8]}"),
                    "detections": len(cam_events),
                    "male_count": cam_male,
                    "female_count": cam_female
                })
        else:
            # If no specific cameras assigned, show all events from this store's VMS
            store_events = [e for e in all_events if e.get("vms_id") == store.get("vms_id")]
        
        # Calculate stats
        total_detections = len(store_events)
        male_count = sum(1 for e in store_events if e.get("gender") == "Male")
        female_count = sum(1 for e in store_events if e.get("gender") == "Female")
        unknown_count = total_detections - male_count - female_count
        
        # Age distribution
        age_dist = {"0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0}
        for event in store_events:
            age = event.get("age", 0)
            if isinstance(age, str):
                try:
                    age = int(age)
                except:
                    age = 0
            if age < 18:
                age_dist["0-17"] += 1
            elif age < 25:
                age_dist["18-24"] += 1
            elif age < 35:
                age_dist["25-34"] += 1
            elif age < 45:
                age_dist["35-44"] += 1
            elif age < 55:
                age_dist["45-54"] += 1
            else:
                age_dist["55+"] += 1
        
        result.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "district_id": store.get("district_id"),
            "district_name": district["name"] if district else "",
            "city_id": city["id"] if city else "",
            "city_name": city["name"] if city else "",
            "region_id": region["id"] if region else "",
            "region_name": region["name"] if region else "",
            "total_detections": total_detections,
            "male_count": male_count,
            "female_count": female_count,
            "unknown_count": unknown_count,
            "male_percent": round(male_count / total_detections * 100, 1) if total_detections > 0 else 0,
            "female_percent": round(female_count / total_detections * 100, 1) if total_detections > 0 else 0,
            "age_distribution": age_dist,
            "camera_count": len(analytics_camera_ids),
            "camera_details": camera_details,
            "events": store_events[:10]  # Last 10 events
        })
    
    return result

# ============== REPORT ENDPOINTS ==============

class ReportRequest(BaseModel):
    report_type: str  # counter, queue, analytics, all
    date_range: str  # 1d, 1w, 1m, 1y, custom
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    region_id: Optional[str] = None
    city_id: Optional[str] = None
    district_id: Optional[str] = None
    store_ids: Optional[List[str]] = None

@api_router.get("/reports/summary")
async def get_summary_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    allowed_stores: List[str] = None
):
    """Get summary report data"""
    
    # Apply permission filtering to store_ids
    if store_ids:
        requested_ids = store_ids.split(",")
        if allowed_stores is not None:
            requested_ids = [sid for sid in requested_ids if sid in allowed_stores]
        filtered_store_ids = requested_ids if requested_ids else []
    elif allowed_stores is not None:
        if not allowed_stores:
            return {
                "counter_summary": {"total_stores": 0, "total_visitors": 0, "total_in": 0, "total_out": 0, "stores_critical": 0, "stores_warning": 0, "stores_normal": 0},
                "queue_summary": {"total_stores": 0, "total_queue_length": 0, "stores_critical": 0, "stores_warning": 0, "stores_normal": 0},
                "analytics_summary": {"total_events": 0, "gender_distribution": {}, "age_distribution": {}, "events": []}
            }
        filtered_store_ids = list(allowed_stores)
    else:
        filtered_store_ids = store_ids.split(",") if store_ids else None
    
    # Get stores directly with permission filter
    store_query = {}
    if filtered_store_ids:
        store_query["id"] = {"$in": filtered_store_ids}
    stores = await db.stores.find(store_query, {"_id": 0}).to_list(100)
    
    # Build counter data
    counter_data = []
    vms_servers = await db.vms_servers.find({"is_active": True}, {"_id": 0}).to_list(100)
    vms_dict = {v["id"]: v for v in vms_servers}
    
    vms_data = {}
    for vms_id, vms in vms_dict.items():
        xml_data = await fetch_vms_data(vms, "/rsapi/modules/counter/getstats")
        if xml_data:
            parsed = parse_counter_xml(xml_data)
            # parsed is {'cameras': [...]} dict
            for p in parsed.get('cameras', []):
                in_count = sum(c.get('in_count', 0) for c in p.get('counters', []))
                out_count = sum(c.get('out_count', 0) for c in p.get('counters', []))
                vms_data[p["camera_id"]] = {
                    "camera_id": p["camera_id"],
                    "in_count": in_count,
                    "out_count": out_count
                }
    
    for store in stores:
        counter_camera_ids = store.get("counter_camera_ids", [])
        if store.get("counter_camera_id") and store["counter_camera_id"] not in counter_camera_ids:
            counter_camera_ids.append(store["counter_camera_id"])
        
        total_in = 0
        total_out = 0
        for cam_id in counter_camera_ids:
            cam_data = vms_data.get(cam_id)
            if cam_data:
                total_in += cam_data.get("in_count", 0)
                total_out += cam_data.get("out_count", 0)
        
        current_visitors = max(0, total_in - total_out)
        occupancy_percent = (current_visitors / store["capacity"] * 100) if store.get("capacity", 0) > 0 else 0
        
        district = await db.districts.find_one({"id": store.get("district_id")}, {"_id": 0})
        city = None
        region = None
        if district:
            city = await db.cities.find_one({"id": district.get("city_id")}, {"_id": 0})
            if city:
                region = await db.regions.find_one({"id": city.get("region_id")}, {"_id": 0})
        
        counter_data.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "district_id": store.get("district_id"),
            "district_name": district["name"] if district else "",
            "city_id": city["id"] if city else "",
            "city_name": city["name"] if city else "",
            "region_id": region["id"] if region else "",
            "region_name": region["name"] if region else "",
            "total_in": total_in,
            "total_out": total_out,
            "current_visitors": current_visitors,
            "capacity": store.get("capacity", 0),
            "occupancy_percent": round(occupancy_percent, 1),
            "status": "critical" if occupancy_percent >= 95 else "warning" if occupancy_percent >= 80 else "normal"
        })
    
    # Build queue data similarly
    queue_vms_data = {}
    for vms_id, vms in vms_dict.items():
        xml_data = await fetch_vms_data(vms, "/rsapi/modules/queue/getstats")
        if xml_data:
            parsed = parse_queue_xml(xml_data)
            # parsed is {'cameras': [...]} dict
            for p in parsed.get('cameras', []):
                queue_vms_data[p["camera_id"]] = p
    
    queue_data = []
    for store in stores:
        queue_camera_ids = store.get("queue_camera_ids", [])
        if store.get("queue_camera_id") and store["queue_camera_id"] not in queue_camera_ids:
            queue_camera_ids.append(store["queue_camera_id"])
        
        total_queue = 0
        for cam_id in queue_camera_ids:
            cam_data = queue_vms_data.get(cam_id)
            if cam_data:
                for zone in cam_data.get("zones", []):
                    total_queue += zone.get("queue_length", 0)
        
        queue_data.append({
            "store_id": store["id"],
            "store_name": store["name"],
            "total_queue_length": total_queue,
            "queue_threshold": store.get("queue_threshold", 5),
            "status": "critical" if total_queue >= store.get("queue_threshold", 5) * 2 else "warning" if total_queue >= store.get("queue_threshold", 5) else "normal"
        })
    
    # Filter by location if needed
    if region_id:
        counter_data = [d for d in counter_data if d.get("region_id") == region_id]
        queue_data = [d for d in queue_data if d.get("region_id") == region_id]
    if city_id:
        counter_data = [d for d in counter_data if d.get("city_id") == city_id]
        queue_data = [d for d in queue_data if d.get("city_id") == city_id]
    if district_id:
        counter_data = [d for d in counter_data if d.get("district_id") == district_id]
        queue_data = [d for d in queue_data if d.get("district_id") == district_id]
    
    analytics_data = {"total_events": 0, "gender_distribution": {}, "age_distribution": {}, "events": []}
    
    return {
        "counter_summary": {
            "total_stores": len(counter_data),
            "total_visitors": sum(d["current_visitors"] for d in counter_data),
            "total_in": sum(d["total_in"] for d in counter_data),
            "total_out": sum(d["total_out"] for d in counter_data),
            "stores_critical": len([d for d in counter_data if d["status"] == "critical"]),
            "stores_warning": len([d for d in counter_data if d["status"] == "warning"]),
            "stores_normal": len([d for d in counter_data if d["status"] == "normal"])
        },
        "queue_summary": {
            "total_stores": len(queue_data),
            "total_queue_length": sum(d["total_queue_length"] for d in queue_data),
            "stores_critical": len([d for d in queue_data if d["status"] == "critical"]),
            "stores_warning": len([d for d in queue_data if d["status"] == "warning"]),
            "stores_normal": len([d for d in queue_data if d["status"] == "normal"])
        },
        "analytics_summary": analytics_data,
        "stores": counter_data,
        "queues": queue_data
    }

@api_router.get("/reports/counter")
async def get_counter_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    hour_from: Optional[int] = None,
    hour_to: Optional[int] = None,
    allowed_stores: List[str] = None
):
    """Get people counter report from local database (cumulative data)"""
    from datetime import timedelta
    
    logger.info(f"Counter report: date_range={date_range}, date_from={date_from}, date_to={date_to}, hour_from={hour_from}, hour_to={hour_to}")
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    if date_from and date_to:
        start_date = date_from
        end_date = date_to
    elif date_range == "1d":
        start_date = end_date = today
    elif date_range in ("1w", "7d"):  # Support both formats
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = today
    elif date_range in ("1m", "30d"):  # Support both formats
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        end_date = today
    else:
        start_date = end_date = today
    
    # Get filtered store IDs
    filtered_ids = None
    if store_ids:
        filtered_ids = store_ids.split(",")
        if allowed_stores is not None:
            filtered_ids = [sid for sid in filtered_ids if sid in allowed_stores]
    elif allowed_stores is not None:
        filtered_ids = list(allowed_stores)
    
    # Build query
    query = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_ids:
        query["store_id"] = {"$in": filtered_ids}
    
    # Add hour filter if specified
    if hour_from is not None:
        query["hour"] = {"$gte": hour_from}
    if hour_to is not None:
        if "hour" in query:
            query["hour"]["$lte"] = hour_to
        else:
            query["hour"] = {"$lte": hour_to}
    
    counter_data = []
    
    # Get store details with location info
    stores_map = {}
    stores = await db.stores.find({}, {"_id": 0}).to_list(500)
    districts = await db.districts.find({}, {"_id": 0}).to_list(500)
    cities = await db.cities.find({}, {"_id": 0}).to_list(500)
    regions = await db.regions.find({}, {"_id": 0}).to_list(500)
    
    districts_map = {d["id"]: d for d in districts}
    cities_map = {c["id"]: c for c in cities}
    regions_map = {r["id"]: r for r in regions}
    
    for s in stores:
        district = districts_map.get(s.get("district_id"), {})
        city = cities_map.get(district.get("city_id"), {})
        region = regions_map.get(city.get("region_id"), {})
        stores_map[s["id"]] = {
            **s,
            "district_name": district.get("name", ""),
            "city_id": district.get("city_id", ""),
            "city_name": city.get("name", ""),
            "region_id": city.get("region_id", ""),
            "region_name": region.get("name", "")
        }
    
    logger.info(f"Counter report query: {query}")
    
    # For today only (without hour filter), use latest snapshots (real-time)
    if start_date == today and end_date == today and hour_from is None and hour_to is None:
        # Get latest snapshot for each store
        pipeline = [
            {"$match": query},
            {"$sort": {"hour": -1, "minute": -1}},
            {"$group": {
                "_id": "$store_id",
                "latest": {"$first": "$$ROOT"}
            }}
        ]
        results = await db.counter_snapshots.aggregate(pipeline).to_list(500)
        
        for r in results:
            snap = r["latest"]
            store = stores_map.get(snap["store_id"], {})
            counter_data.append({
                "store_id": snap["store_id"],
                "store_name": store.get("name", snap.get("store_name", "Bilinmiyor")),
                "region_id": store.get("region_id", ""),
                "region_name": store.get("region_name", ""),
                "city_id": store.get("city_id", ""),
                "city_name": store.get("city_name", ""),
                "district_id": store.get("district_id", ""),
                "district_name": store.get("district_name", ""),
                "total_in": snap.get("total_in", 0),
                "total_out": snap.get("total_out", 0),
                "current_visitors": snap.get("current_visitors", 0),
                "capacity": store.get("capacity", 100),
                "occupancy_percent": round(snap.get("current_visitors", 0) / store.get("capacity", 100) * 100, 1) if store.get("capacity", 100) > 0 else 0,
                "status": snap.get("status", "normal")
            })
    else:
        # For historical data OR hour-filtered data
        # When hour filter is applied, we need to aggregate snapshots within that hour range
        if hour_from is not None or hour_to is not None:
            # Hour filter: aggregate all snapshots in the hour range
            # Get the MAX total_in for each store in the filtered period (cumulative at end of period)
            pipeline = [
                {"$match": query},
                {"$sort": {"date": 1, "hour": 1, "minute": 1}},  # Sort ascending to get progression
                {"$group": {
                    "_id": "$store_id",
                    "max_in": {"$max": "$total_in"},
                    "max_out": {"$max": "$total_out"},
                    "last_visitors": {"$last": "$current_visitors"},
                    "store_name": {"$first": "$store_name"},
                    "snapshot_count": {"$sum": 1}
                }}
            ]
            results = await db.counter_snapshots.aggregate(pipeline).to_list(500)
            
            logger.info(f"Hour-filtered report: Found {len(results)} stores with {sum(r.get('snapshot_count', 0) for r in results)} total snapshots")
            
            for r in results:
                store = stores_map.get(r["_id"], {})
                counter_data.append({
                    "store_id": r["_id"],
                    "store_name": store.get("name", r.get("store_name", "Bilinmiyor")),
                    "region_id": store.get("region_id", ""),
                    "region_name": store.get("region_name", ""),
                    "city_id": store.get("city_id", ""),
                    "city_name": store.get("city_name", ""),
                    "district_id": store.get("district_id", ""),
                    "district_name": store.get("district_name", ""),
                    "total_in": r.get("max_in", 0),
                    "total_out": r.get("max_out", 0),
                    "current_visitors": r.get("last_visitors", 0),
                    "capacity": store.get("capacity", 100),
                    "occupancy_percent": round(r.get("last_visitors", 0) / store.get("capacity", 100) * 100, 1) if store.get("capacity", 100) > 0 else 0,
                    "status": "normal"
                })
        else:
            # No hour filter: Get the LAST snapshot of each day for each store (cumulative value at end of day)
            pipeline = [
                {"$match": query},
                {"$sort": {"hour": -1, "minute": -1}},  # Sort by time descending
                {"$group": {
                    "_id": {"store_id": "$store_id", "date": "$date"},
                    "latest": {"$first": "$$ROOT"}
                }},
                {"$replaceRoot": {"newRoot": "$latest"}},
                {"$project": {"_id": 0}}
            ]
            snapshots = await db.counter_snapshots.aggregate(pipeline).to_list(10000)
            
            logger.info(f"Historical report: Found {len(snapshots)} daily snapshots for date range {start_date} to {end_date}")
            
            # Aggregate by store across all days
            store_data = {}
            for snap in snapshots:
                sid = snap["store_id"]
                if sid not in store_data:
                    store = stores_map.get(sid, {})
                    store_data[sid] = {
                        "store_id": sid,
                        "store_name": store.get("name", snap.get("store_name", "Bilinmiyor")),
                        "region_id": store.get("region_id", ""),
                        "region_name": store.get("region_name", ""),
                        "city_id": store.get("city_id", ""),
                        "city_name": store.get("city_name", ""),
                        "district_id": store.get("district_id", ""),
                        "district_name": store.get("district_name", ""),
                        "total_in": 0,
                        "total_out": 0,
                        "max_visitors": 0,
                        "capacity": store.get("capacity", 100),
                        "days": 0
                    }
                # Sum each day's final values (VMS resets daily)
                store_data[sid]["total_in"] += snap.get("total_in", 0)
                store_data[sid]["total_out"] += snap.get("total_out", 0)
                store_data[sid]["max_visitors"] = max(store_data[sid]["max_visitors"], snap.get("current_visitors", 0))
                store_data[sid]["days"] += 1
            
            for sid, data in store_data.items():
                data["current_visitors"] = max(0, data["total_in"] - data["total_out"])
                data["occupancy_percent"] = round(data["current_visitors"] / data["capacity"] * 100, 1) if data["capacity"] > 0 else 0
                data["status"] = "normal" if data["occupancy_percent"] < 70 else ("warning" if data["occupancy_percent"] < 90 else "critical")
                counter_data.append(data)
            
            logger.info(f"Historical report: Aggregated to {len(counter_data)} stores with total_in={sum(d['total_in'] for d in counter_data)}")
    
    # Filter by location
    if region_id:
        counter_data = [d for d in counter_data if d.get("region_id") == region_id]
    if city_id:
        counter_data = [d for d in counter_data if d.get("city_id") == city_id]
    if district_id:
        counter_data = [d for d in counter_data if d.get("district_id") == district_id]
    
    # Calculate totals
    total_in = sum(d["total_in"] for d in counter_data)
    total_out = sum(d["total_out"] for d in counter_data)
    current_visitors = sum(d["current_visitors"] for d in counter_data)
    avg_occupancy = sum(d["occupancy_percent"] for d in counter_data) / len(counter_data) if counter_data else 0
    
    return {
        "report_type": "counter",
        "date_range": date_range,
        "date_from": start_date,
        "date_to": end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_stores": len(counter_data),
            "total_in": total_in,
            "total_out": total_out,
            "current_visitors": current_visitors,
            "avg_occupancy": round(avg_occupancy, 1),
            "stores_critical": len([d for d in counter_data if d["status"] == "critical"]),
            "stores_warning": len([d for d in counter_data if d["status"] == "warning"]),
            "stores_normal": len([d for d in counter_data if d["status"] == "normal"])
        },
        "stores": counter_data
    }

@api_router.get("/reports/queue")
async def get_queue_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    hour_from: Optional[int] = None,
    hour_to: Optional[int] = None,
    allowed_stores: List[str] = None
):
    """Get queue analysis report from snapshots or live data
    
    FIXED: Now shows meaningful metrics:
    - avg_queue_length: Average queue length across all measurements
    - max_queue_length: Maximum queue length observed
    - threshold_exceed_count: Number of times threshold was exceeded
    """
    from datetime import timedelta
    
    logger.info(f"Queue report: date_range={date_range}, date_from={date_from}, date_to={date_to}, hour_from={hour_from}, hour_to={hour_to}")
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    if date_from and date_to:
        start_date = date_from
        end_date = date_to
    elif date_range == "1d":
        start_date = end_date = today
    elif date_range in ("1w", "7d"):  # Support both formats
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = today
    elif date_range in ("1m", "30d"):  # Support both formats
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        end_date = today
    else:
        start_date = end_date = today
    
    # Get store location info
    stores = await db.stores.find({}, {"_id": 0}).to_list(500)
    districts = await db.districts.find({}, {"_id": 0}).to_list(500)
    cities = await db.cities.find({}, {"_id": 0}).to_list(500)
    regions = await db.regions.find({}, {"_id": 0}).to_list(500)
    
    districts_map = {d["id"]: d for d in districts}
    cities_map = {c["id"]: c for c in cities}
    regions_map = {r["id"]: r for r in regions}
    
    stores_map = {}
    for s in stores:
        district = districts_map.get(s.get("district_id"), {})
        city = cities_map.get(district.get("city_id"), {})
        region = regions_map.get(city.get("region_id"), {})
        stores_map[s["id"]] = {
            **s,
            "district_name": district.get("name", ""),
            "city_name": city.get("name", ""),
            "region_name": region.get("name", "")
        }
    
    queue_data = []
    
    # For today, use live data
    if start_date == today and end_date == today and hour_from is None and hour_to is None:
        live_data = await _fetch_live_queue_data(store_ids, allowed_stores)
        for d in live_data:
            store = stores_map.get(d["store_id"], {})
            d["region_name"] = store.get("region_name", "")
            d["city_name"] = store.get("city_name", "")
            d["district_name"] = store.get("district_name", "")
            # Add default values for new fields in live data
            d["avg_queue_length"] = d.get("total_queue_length", 0)
            d["max_queue_length"] = d.get("total_queue_length", 0)
            d["threshold_exceed_count"] = 1 if d.get("total_queue_length", 0) >= d.get("queue_threshold", 5) else 0
            d["measurement_count"] = 1
        queue_data = live_data
    else:
        # For historical data, use queue_snapshots with MEANINGFUL aggregations
        query = {"date": {"$gte": start_date, "$lte": end_date}}
        
        # Add hour filter if specified
        if hour_from is not None:
            query["hour"] = {"$gte": hour_from}
        if hour_to is not None:
            if "hour" in query:
                query["hour"]["$lte"] = hour_to
            else:
                query["hour"] = {"$lte": hour_to}
        
        filtered_ids = None
        if store_ids:
            filtered_ids = store_ids.split(",")
            if allowed_stores is not None:
                filtered_ids = [sid for sid in filtered_ids if sid in allowed_stores]
            query["store_id"] = {"$in": filtered_ids}
        elif allowed_stores is not None:
            query["store_id"] = {"$in": list(allowed_stores)}
        
        logger.info(f"Queue report query: {query}")
        
        # FIXED: Calculate AVERAGE, MAX, and THRESHOLD EXCEED COUNT instead of SUM
        pipeline = [
            {"$match": query},
            {"$group": {
                "_id": "$store_id",
                "avg_queue": {"$avg": "$total_queue_length"},  # Average queue length
                "max_queue": {"$max": "$total_queue_length"},  # Maximum queue length
                "snapshot_count": {"$sum": 1},
                "queue_threshold": {"$first": "$queue_threshold"}  # Store threshold
            }}
        ]
        results = await db.queue_snapshots.aggregate(pipeline).to_list(500)
        
        # Now calculate threshold exceed count with a separate query per store
        for r in results:
            sid = r["_id"]
            if not sid:  # Skip entries without store_id (this fixes "Bilinmiyor")
                continue
                
            store = stores_map.get(sid)
            if not store:  # Skip if store not found in stores collection
                continue
            
            # Get the threshold for this store
            threshold = store.get("queue_threshold", 5)
            
            # Count how many times threshold was exceeded
            exceed_count = await db.queue_snapshots.count_documents({
                **query,
                "store_id": sid,
                "total_queue_length": {"$gte": threshold}
            })
            
            avg_queue = round(r["avg_queue"] or 0, 1)
            max_queue = r["max_queue"] or 0
            
            # Determine status based on AVERAGE queue, not sum
            if avg_queue >= threshold * 1.5:
                status = "critical"
            elif avg_queue >= threshold:
                status = "warning"
            else:
                status = "normal"
            
            queue_data.append({
                "store_id": sid,
                "store_name": store.get("name", ""),
                "region_id": store.get("region_id", ""),
                "region_name": store.get("region_name", ""),
                "city_id": store.get("city_id", ""),
                "city_name": store.get("city_name", ""),
                "district_id": store.get("district_id", ""),
                "district_name": store.get("district_name", ""),
                "avg_queue_length": avg_queue,  # MEANINGFUL: Average queue
                "max_queue_length": max_queue,  # MEANINGFUL: Max queue seen
                "threshold_exceed_count": exceed_count,  # MEANINGFUL: How many times threshold exceeded
                "measurement_count": r["snapshot_count"],  # How many measurements
                "queue_threshold": threshold,
                "status": status,
                # Keep total_queue_length for backward compatibility but mark as deprecated
                "total_queue_length": avg_queue  # Use avg for display
            })
    
    # Filter by location
    if region_id:
        queue_data = [d for d in queue_data if d.get("region_id") == region_id]
    if city_id:
        queue_data = [d for d in queue_data if d.get("city_id") == city_id]
    if district_id:
        queue_data = [d for d in queue_data if d.get("district_id") == district_id]
    
    # Calculate summary with meaningful metrics
    total_avg_queue = sum(d.get("avg_queue_length", 0) for d in queue_data)
    overall_avg = round(total_avg_queue / len(queue_data), 1) if queue_data else 0
    total_max_queue = max((d.get("max_queue_length", 0) for d in queue_data), default=0)
    total_exceed_count = sum(d.get("threshold_exceed_count", 0) for d in queue_data)
    
    return {
        "report_type": "queue",
        "date_range": date_range,
        "date_from": date_from or start_date,
        "date_to": date_to or end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_stores": len(queue_data),
            "avg_queue_overall": overall_avg,  # Average across all stores
            "max_queue_observed": total_max_queue,  # Highest queue seen
            "total_threshold_exceeds": total_exceed_count,  # Total exceed events
            "stores_critical": len([d for d in queue_data if d.get("status") == "critical"]),
            "stores_warning": len([d for d in queue_data if d.get("status") == "warning"]),
            "stores_normal": len([d for d in queue_data if d.get("status") == "normal"]),
            # Backward compatibility
            "total_queue_length": round(overall_avg * len(queue_data), 0),
            "avg_queue_per_store": overall_avg
        },
        "stores": queue_data
    }

@api_router.get("/reports/analytics")
async def get_analytics_report(
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    hour_from: Optional[int] = None,
    hour_to: Optional[int] = None
):
    """Get age/gender analytics report with store details"""
    from datetime import timedelta
    import logging
    log = logging.getLogger(__name__)
    
    log.info(f"Analytics report: date_range={date_range}, date_from={date_from}, date_to={date_to}, hour_from={hour_from}, hour_to={hour_to}")
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    
    if date_from and date_to:
        start_date = date_from
        end_date = date_to
    elif date_range == "1d":
        start_date = end_date = today
    elif date_range in ("1w", "7d"):  # Support both formats
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = today
    elif date_range in ("1m", "30d"):  # Support both formats
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        end_date = today
    else:
        start_date = end_date = today
    
    # Get store location info
    stores = await db.stores.find({}, {"_id": 0}).to_list(500)
    districts = await db.districts.find({}, {"_id": 0}).to_list(500)
    cities = await db.cities.find({}, {"_id": 0}).to_list(500)
    regions = await db.regions.find({}, {"_id": 0}).to_list(500)
    
    districts_map = {d["id"]: d for d in districts}
    cities_map = {c["id"]: c for c in cities}
    regions_map = {r["id"]: r for r in regions}
    
    stores_map = {}
    for s in stores:
        district = districts_map.get(s.get("district_id"), {})
        city = cities_map.get(district.get("city_id"), {})
        region = regions_map.get(city.get("region_id"), {})
        stores_map[s["id"]] = {
            **s,
            "district_name": district.get("name", ""),
            "city_id": district.get("city_id", ""),
            "city_name": city.get("name", ""),
            "region_id": city.get("region_id", ""),
            "region_name": region.get("name", "")
        }
    
    # Filter store IDs
    filtered_ids = None
    if store_ids:
        filtered_ids = store_ids.split(",")
    
    # Build query for analytics_snapshots
    query = {"date": {"$gte": start_date, "$lte": end_date}}
    if filtered_ids:
        query["store_id"] = {"$in": filtered_ids}
    
    # Add hour filter if specified
    if hour_from is not None:
        query["hour"] = {"$gte": hour_from}
    if hour_to is not None:
        if "hour" in query:
            query["hour"]["$lte"] = hour_to
        else:
            query["hour"] = {"$lte": hour_to}
    
    logger.info(f"Analytics report query: {query}")
    
    # Get analytics data from database (aggregated snapshots)
    snapshots = await db.analytics_snapshots.find(query, {"_id": 0}).to_list(10000)
    
    # Check if snapshots have meaningful data
    total_snapshot_events = sum(s.get("total_events", 0) for s in snapshots)
    has_snapshot_data = len(snapshots) > 0 and total_snapshot_events > 0
    
    # IMPORTANT: Log the data source being used for debugging
    log.info(f"Analytics Report: date_range={date_range}, start={start_date}, end={end_date}, snapshots={len(snapshots)}, total_events={total_snapshot_events}, using_snapshots={has_snapshot_data}")
    
    # Aggregate by store
    store_analytics = {}
    total_gender = {"Male": 0, "Female": 0, "Unknown": 0}
    total_age = {"0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0}
    
    # Data source indicator for transparency
    data_source = "snapshots" if has_snapshot_data else "live_vms"
    
    if has_snapshot_data:
        # Use snapshots data
        for snap in snapshots:
            sid = snap["store_id"]
            if sid not in store_analytics:
                store = stores_map.get(sid, {})
                store_analytics[sid] = {
                    "store_id": sid,
                    "store_name": store.get("name", snap.get("store_name", "Bilinmiyor")),
                    "region_id": store.get("region_id", ""),
                    "region_name": store.get("region_name", ""),
                    "city_id": store.get("city_id", ""),
                    "city_name": store.get("city_name", ""),
                    "district_id": store.get("district_id", ""),
                    "district_name": store.get("district_name", ""),
                    "total_detections": 0,
                    "male_count": 0,
                    "female_count": 0,
                    "age_distribution": {"0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0}
                }
            
            gender = snap.get("gender_distribution", {})
            age = snap.get("age_distribution", {})
            
            store_analytics[sid]["total_detections"] += snap.get("total_events", 0)
            store_analytics[sid]["male_count"] += gender.get("Male", 0)
            store_analytics[sid]["female_count"] += gender.get("Female", 0)
            
            for age_group in total_age.keys():
                store_analytics[sid]["age_distribution"][age_group] += age.get(age_group, 0)
                total_age[age_group] += age.get(age_group, 0)
            
            total_gender["Male"] += gender.get("Male", 0)
            total_gender["Female"] += gender.get("Female", 0)
    else:
        # Fallback: Use live VMS data
        # Get analytics camera IDs for stores
        target_stores = list(stores_map.values())
        if filtered_ids:
            target_stores = [s for s in target_stores if s["id"] in filtered_ids]
        
        for store in target_stores:
            sid = store["id"]
            analytics_camera_ids = store.get("analytics_camera_ids", [])
            if store.get("analytics_camera_id") and store["analytics_camera_id"] not in analytics_camera_ids:
                analytics_camera_ids.append(store["analytics_camera_id"])
            
            # Get live data from VMS
            live_data = await _fetch_analytics_data(
                store_ids=sid,
                allowed_camera_ids=analytics_camera_ids if analytics_camera_ids else None
            )
            
            gender = live_data.get("gender_distribution", {})
            age = live_data.get("age_distribution", {})
            total_events = live_data.get("total_events", 0)
            
            store_analytics[sid] = {
                "store_id": sid,
                "store_name": store.get("name", "Bilinmiyor"),
                "region_id": store.get("region_id", ""),
                "region_name": store.get("region_name", ""),
                "city_id": store.get("city_id", ""),
                "city_name": store.get("city_name", ""),
                "district_id": store.get("district_id", ""),
                "district_name": store.get("district_name", ""),
                "total_detections": total_events,
                "male_count": gender.get("Male", 0),
                "female_count": gender.get("Female", 0),
                "age_distribution": age.copy()
            }
            
            total_gender["Male"] += gender.get("Male", 0)
            total_gender["Female"] += gender.get("Female", 0)
            for age_group in total_age.keys():
                total_age[age_group] += age.get(age_group, 0)
    
    # Calculate percentages for each store
    stores_list = []
    for sid, data in store_analytics.items():
        total = data["total_detections"]
        data["male_percent"] = round(data["male_count"] / total * 100, 1) if total > 0 else 0
        data["female_percent"] = round(data["female_count"] / total * 100, 1) if total > 0 else 0
        stores_list.append(data)
    
    # Filter by location
    if region_id:
        stores_list = [d for d in stores_list if d.get("region_id") == region_id]
    if city_id:
        stores_list = [d for d in stores_list if d.get("city_id") == city_id]
    if district_id:
        stores_list = [d for d in stores_list if d.get("district_id") == district_id]
    
    # Recalculate totals after filtering
    total_detections = sum(s["total_detections"] for s in stores_list)
    total_male = sum(s["male_count"] for s in stores_list)
    total_female = sum(s["female_count"] for s in stores_list)
    
    # Recalculate age totals
    final_age = {"0-17": 0, "18-24": 0, "25-34": 0, "35-44": 0, "45-54": 0, "55+": 0}
    for s in stores_list:
        for age_group in final_age.keys():
            final_age[age_group] += s["age_distribution"].get(age_group, 0)
    
    return {
        "report_type": "analytics",
        "date_range": date_range,
        "date_from": date_from or start_date,
        "date_to": date_to or end_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "data_source": data_source,  # "snapshots" or "live_vms" - for transparency
        "data_source_note": "Geçmiş veriler (snapshots)" if data_source == "snapshots" else "Canlı VMS verisi (anlık) - geçmiş veri bulunamadı",
        "summary": {
            "total_stores": len(stores_list),
            "total_detections": total_detections,
            "male_count": total_male,
            "female_count": total_female,
            "male_percent": round(total_male / total_detections * 100, 1) if total_detections > 0 else 0,
            "female_percent": round(total_female / total_detections * 100, 1) if total_detections > 0 else 0
        },
        "gender_distribution": {"Male": total_male, "Female": total_female},
        "age_distribution": final_age,
        "stores": stores_list
    }

@api_router.get("/reports/export")
async def export_report(
    report_type: str = "counter",
    format: str = "json",
    region_id: Optional[str] = None,
    city_id: Optional[str] = None,
    district_id: Optional[str] = None,
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Export report as JSON, CSV or Excel"""
    import xlsxwriter
    from permissions import get_user_allowed_stores
    
    # Get user's allowed stores for permission filtering
    allowed_stores = await get_user_allowed_stores(user)
    
    # Get report data based on type
    if report_type == "counter":
        report = await get_counter_report(region_id, city_id, district_id, store_ids, date_range, date_from, date_to, allowed_stores)
    elif report_type == "queue":
        report = await get_queue_report(region_id, city_id, district_id, store_ids, date_range, date_from, date_to, allowed_stores)
    elif report_type == "analytics":
        report = await get_analytics_report(region_id, city_id, district_id, store_ids, date_range, date_from, date_to)
    else:
        report = await get_summary_report(region_id, city_id, district_id, store_ids, allowed_stores)
    
    if format == "json":
        return report
    
    elif format == "csv":
        csv_lines = []
        
        if report_type == "counter":
            csv_lines.append("Mağaza,Bölge,İl,İlçe,Giriş,Çıkış,Mevcut,Kapasite,Doluluk %,Durum")
            for store in report.get("stores", []):
                csv_lines.append(
                    f"{store['store_name']},{store.get('region_name','')},{store.get('city_name','')},{store.get('district_name','')},"
                    f"{store['total_in']},{store['total_out']},{store['current_visitors']},"
                    f"{store['capacity']},{store['occupancy_percent']},{store['status']}"
                )
        elif report_type == "queue":
            csv_lines.append("Mağaza,Bölge,İl,İlçe,Ort. Kuyruk,Maks. Kuyruk,Eşik Aşım,Ölçüm Sayısı,Eşik,Durum")
            for store in report.get("stores", []):
                csv_lines.append(
                    f"{store['store_name']},{store.get('region_name','')},{store.get('city_name','')},{store.get('district_name','')},"
                    f"{store.get('avg_queue_length',0)},{store.get('max_queue_length',0)},{store.get('threshold_exceed_count',0)},"
                    f"{store.get('measurement_count',0)},{store.get('queue_threshold',5)},{store['status']}"
                )
        elif report_type == "analytics":
            csv_lines.append("Zaman,Kamera,Yaş,Cinsiyet,Tanınma")
            for event in report.get("events", []):
                csv_lines.append(
                    f"{event.get('time','')},{event.get('camera_name','')},{event.get('age','')},{event.get('gender','')},{event.get('is_recognized','')}"
                )
        
        csv_content = "\n".join(csv_lines)
        filename = f"rapor_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    elif format == "excel":
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        
        # Styles
        header_format = workbook.add_format({'bold': True, 'bg_color': '#3B82F6', 'font_color': 'white', 'border': 1})
        cell_format = workbook.add_format({'border': 1})
        number_format = workbook.add_format({'border': 1, 'num_format': '#,##0'})
        percent_format = workbook.add_format({'border': 1, 'num_format': '0.0%'})
        status_normal = workbook.add_format({'border': 1, 'bg_color': '#D1FAE5', 'font_color': '#065F46'})
        status_warning = workbook.add_format({'border': 1, 'bg_color': '#FEF3C7', 'font_color': '#92400E'})
        status_critical = workbook.add_format({'border': 1, 'bg_color': '#FEE2E2', 'font_color': '#991B1B'})
        
        if report_type == "counter":
            # Summary sheet
            summary_sheet = workbook.add_worksheet("Özet")
            summary_sheet.write(0, 0, "Kişi Sayma Raporu", header_format)
            summary_sheet.write(1, 0, f"Tarih Aralığı: {date_range}", cell_format)
            summary_sheet.write(2, 0, f"Oluşturulma: {datetime.now().strftime('%Y-%m-%d %H:%M')}", cell_format)
            
            summary = report.get("summary", {})
            summary_sheet.write(4, 0, "Toplam Mağaza", cell_format)
            summary_sheet.write(4, 1, summary.get("total_stores", 0), number_format)
            summary_sheet.write(5, 0, "Toplam Giriş", cell_format)
            summary_sheet.write(5, 1, summary.get("total_in", 0), number_format)
            summary_sheet.write(6, 0, "Toplam Çıkış", cell_format)
            summary_sheet.write(6, 1, summary.get("total_out", 0), number_format)
            summary_sheet.write(7, 0, "Mevcut Ziyaretçi", cell_format)
            summary_sheet.write(7, 1, summary.get("current_visitors", 0), number_format)
            summary_sheet.write(8, 0, "Ort. Doluluk", cell_format)
            summary_sheet.write(8, 1, summary.get("avg_occupancy", 0) / 100, percent_format)
            
            # Detail sheet
            detail_sheet = workbook.add_worksheet("Mağaza Detayları")
            headers = ["Mağaza", "Bölge", "İl", "İlçe", "Giriş", "Çıkış", "Mevcut", "Kapasite", "Doluluk %", "Durum"]
            for col, header in enumerate(headers):
                detail_sheet.write(0, col, header, header_format)
            
            for row, store in enumerate(report.get("stores", []), 1):
                detail_sheet.write(row, 0, store.get("store_name", ""), cell_format)
                detail_sheet.write(row, 1, store.get("region_name", ""), cell_format)
                detail_sheet.write(row, 2, store.get("city_name", ""), cell_format)
                detail_sheet.write(row, 3, store.get("district_name", ""), cell_format)
                detail_sheet.write(row, 4, store.get("total_in", 0), number_format)
                detail_sheet.write(row, 5, store.get("total_out", 0), number_format)
                detail_sheet.write(row, 6, store.get("current_visitors", 0), number_format)
                detail_sheet.write(row, 7, store.get("capacity", 0), number_format)
                detail_sheet.write(row, 8, store.get("occupancy_percent", 0) / 100, percent_format)
                status = store.get("status", "normal")
                status_fmt = status_normal if status == "normal" else status_warning if status == "warning" else status_critical
                detail_sheet.write(row, 9, status.upper(), status_fmt)
            
            # Add TOPLAM (Total) row
            total_row = len(report.get("stores", [])) + 1
            total_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1})
            total_number = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '#,##0'})
            total_percent = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '0.0%'})
            
            detail_sheet.write(total_row, 0, "TOPLAM", total_format)
            detail_sheet.write(total_row, 1, "", total_format)
            detail_sheet.write(total_row, 2, "", total_format)
            detail_sheet.write(total_row, 3, "", total_format)
            detail_sheet.write(total_row, 4, summary.get("total_in", 0), total_number)
            detail_sheet.write(total_row, 5, summary.get("total_out", 0), total_number)
            detail_sheet.write(total_row, 6, summary.get("current_visitors", 0), total_number)
            detail_sheet.write(total_row, 7, sum(s.get("capacity", 0) for s in report.get("stores", [])), total_number)
            detail_sheet.write(total_row, 8, summary.get("avg_occupancy", 0) / 100, total_percent)
            detail_sheet.write(total_row, 9, "", total_format)
            
            detail_sheet.autofit()
        
        elif report_type == "queue":
            summary_sheet = workbook.add_worksheet("Özet")
            summary_sheet.write(0, 0, "Kuyruk Analizi Raporu", header_format)
            
            summary = report.get("summary", {})
            summary_sheet.write(2, 0, "Toplam Mağaza", cell_format)
            summary_sheet.write(2, 1, summary.get("total_stores", 0), number_format)
            summary_sheet.write(3, 0, "Ortalama Kuyruk", cell_format)
            summary_sheet.write(3, 1, summary.get("avg_queue_overall", 0), number_format)
            summary_sheet.write(4, 0, "Maksimum Kuyruk", cell_format)
            summary_sheet.write(4, 1, summary.get("max_queue_observed", 0), number_format)
            summary_sheet.write(5, 0, "Toplam Eşik Aşımı", cell_format)
            summary_sheet.write(5, 1, summary.get("total_threshold_exceeds", 0), number_format)
            
            detail_sheet = workbook.add_worksheet("Mağaza Detayları")
            headers = ["Mağaza", "Bölge", "İl", "İlçe", "Ort. Kuyruk", "Maks. Kuyruk", "Eşik Aşım", "Ölçüm", "Eşik", "Durum"]
            for col, header in enumerate(headers):
                detail_sheet.write(0, col, header, header_format)
            
            for row, store in enumerate(report.get("stores", []), 1):
                detail_sheet.write(row, 0, store.get("store_name", ""), cell_format)
                detail_sheet.write(row, 1, store.get("region_name", ""), cell_format)
                detail_sheet.write(row, 2, store.get("city_name", ""), cell_format)
                detail_sheet.write(row, 3, store.get("district_name", ""), cell_format)
                detail_sheet.write(row, 4, store.get("avg_queue_length", 0), number_format)
                detail_sheet.write(row, 5, store.get("max_queue_length", 0), number_format)
                detail_sheet.write(row, 6, store.get("threshold_exceed_count", 0), number_format)
                detail_sheet.write(row, 7, store.get("measurement_count", 0), number_format)
                detail_sheet.write(row, 8, store.get("queue_threshold", 0), number_format)
                status = store.get("status", "normal")
                status_fmt = status_normal if status == "normal" else status_warning if status == "warning" else status_critical
                detail_sheet.write(row, 9, status.upper(), status_fmt)
            
            # Add TOPLAM (Total) row
            total_row = len(report.get("stores", [])) + 1
            total_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1})
            total_number = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '#,##0'})
            
            detail_sheet.write(total_row, 0, "TOPLAM", total_format)
            detail_sheet.write(total_row, 1, "", total_format)
            detail_sheet.write(total_row, 2, "", total_format)
            detail_sheet.write(total_row, 3, "", total_format)
            detail_sheet.write(total_row, 4, summary.get("avg_queue_overall", 0), total_number)
            detail_sheet.write(total_row, 5, summary.get("max_queue_observed", 0), total_number)
            detail_sheet.write(total_row, 6, summary.get("total_threshold_exceeds", 0), total_number)
            detail_sheet.write(total_row, 7, "", total_format)
            detail_sheet.write(total_row, 8, "", total_format)
            detail_sheet.write(total_row, 9, "", total_format)
            
            detail_sheet.autofit()
        
        elif report_type == "analytics":
            summary_sheet = workbook.add_worksheet("Özet")
            summary_sheet.write(0, 0, "Yaş/Cinsiyet Analizi Raporu", header_format)
            summary_sheet.write(1, 0, f"Tarih Aralığı: {date_range}", cell_format)
            summary_sheet.write(2, 0, f"Oluşturulma: {datetime.now().strftime('%Y-%m-%d %H:%M')}", cell_format)
            
            summary = report.get("summary", {})
            summary_sheet.write(4, 0, "Toplam Mağaza", cell_format)
            summary_sheet.write(4, 1, summary.get("total_stores", 0), number_format)
            summary_sheet.write(5, 0, "Toplam Tespit", cell_format)
            summary_sheet.write(5, 1, summary.get("total_detections", 0), number_format)
            summary_sheet.write(6, 0, "Erkek", cell_format)
            summary_sheet.write(6, 1, summary.get("male_count", 0), number_format)
            summary_sheet.write(7, 0, "Kadın", cell_format)
            summary_sheet.write(7, 1, summary.get("female_count", 0), number_format)
            summary_sheet.write(8, 0, "Erkek %", cell_format)
            summary_sheet.write(8, 1, summary.get("male_percent", 0) / 100, percent_format)
            summary_sheet.write(9, 0, "Kadın %", cell_format)
            summary_sheet.write(9, 1, summary.get("female_percent", 0) / 100, percent_format)
            
            # Store details sheet
            detail_sheet = workbook.add_worksheet("Mağaza Detayları")
            headers = ["Mağaza", "Bölge", "İl", "İlçe", "Tespit", "Erkek", "Kadın", "Erkek %", "Kadın %"]
            for col, header in enumerate(headers):
                detail_sheet.write(0, col, header, header_format)
            
            for row, store in enumerate(report.get("stores", []), 1):
                detail_sheet.write(row, 0, store.get("store_name", ""), cell_format)
                detail_sheet.write(row, 1, store.get("region_name", ""), cell_format)
                detail_sheet.write(row, 2, store.get("city_name", ""), cell_format)
                detail_sheet.write(row, 3, store.get("district_name", ""), cell_format)
                detail_sheet.write(row, 4, store.get("total_detections", 0), number_format)
                detail_sheet.write(row, 5, store.get("male_count", 0), number_format)
                detail_sheet.write(row, 6, store.get("female_count", 0), number_format)
                detail_sheet.write(row, 7, store.get("male_percent", 0) / 100, percent_format)
                detail_sheet.write(row, 8, store.get("female_percent", 0) / 100, percent_format)
            
            # Add TOPLAM (Total) row
            total_row = len(report.get("stores", [])) + 1
            total_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1})
            total_number = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '#,##0'})
            total_percent_fmt = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '0.0%'})
            
            detail_sheet.write(total_row, 0, "TOPLAM", total_format)
            detail_sheet.write(total_row, 1, "", total_format)
            detail_sheet.write(total_row, 2, "", total_format)
            detail_sheet.write(total_row, 3, "", total_format)
            detail_sheet.write(total_row, 4, summary.get("total_detections", 0), total_number)
            detail_sheet.write(total_row, 5, summary.get("male_count", 0), total_number)
            detail_sheet.write(total_row, 6, summary.get("female_count", 0), total_number)
            detail_sheet.write(total_row, 7, summary.get("male_percent", 0) / 100, total_percent_fmt)
            detail_sheet.write(total_row, 8, summary.get("female_percent", 0) / 100, total_percent_fmt)
            
            detail_sheet.autofit()
            
            # Age distribution sheet
            age_sheet = workbook.add_worksheet("Yaş Dağılımı")
            age_sheet.write(0, 0, "Yaş Grubu", header_format)
            age_sheet.write(0, 1, "Sayı", header_format)
            age_sheet.write(0, 2, "Yüzde", header_format)
            
            total_age = sum(report.get("age_distribution", {}).values())
            for row, (age_group, count) in enumerate(report.get("age_distribution", {}).items(), 1):
                age_sheet.write(row, 0, age_group, cell_format)
                age_sheet.write(row, 1, count, number_format)
                pct = count / total_age if total_age > 0 else 0
                age_sheet.write(row, 2, pct, percent_format)
            
            # Total row for age
            age_total_row = len(report.get("age_distribution", {})) + 1
            age_sheet.write(age_total_row, 0, "TOPLAM", total_format)
            age_sheet.write(age_total_row, 1, total_age, total_number)
            age_sheet.write(age_total_row, 2, 1.0, total_percent_fmt)
            
            age_sheet.autofit()
        
        workbook.close()
        output.seek(0)
        
        filename = f"rapor_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

# ============== HIERARCHY ENDPOINT ==============

@api_router.get("/hierarchy")
async def get_full_hierarchy():
    """Get full location hierarchy with stores"""
    regions = await db.regions.find({}, {"_id": 0}).to_list(100)
    cities = await db.cities.find({}, {"_id": 0}).to_list(100)
    districts = await db.districts.find({}, {"_id": 0}).to_list(100)
    stores = await db.stores.find({}, {"_id": 0}).to_list(100)
    
    # Build hierarchy
    result = []
    for region in regions:
        region_data = {
            "id": region["id"],
            "name": region["name"],
            "type": "region",
            "cities": []
        }
        
        for city in [c for c in cities if c.get("region_id") == region["id"]]:
            city_data = {
                "id": city["id"],
                "name": city["name"],
                "type": "city",
                "districts": []
            }
            
            for district in [d for d in districts if d.get("city_id") == city["id"]]:
                district_data = {
                    "id": district["id"],
                    "name": district["name"],
                    "type": "district",
                    "stores": [s for s in stores if s.get("district_id") == district["id"]]
                }
                city_data["districts"].append(district_data)
            
            region_data["cities"].append(city_data)
        
        result.append(region_data)
    
    return result

# ============== ADVANCED ANALYTICS ENDPOINTS ==============

@api_router.get("/reports/advanced/hourly-traffic")
async def get_hourly_traffic_report(
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    user: dict = Depends(require_auth)
):
    """Get hourly traffic pattern for stores"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    counter_data = await _fetch_live_counter_data(store_ids, allowed_stores)
    
    # Generate hourly distribution (simulated based on current data)
    hours = list(range(8, 23))  # 08:00 - 22:00
    peak_hours = [11, 12, 13, 17, 18, 19]  # Typical retail peak hours
    
    total_in = sum(s["total_in"] for s in counter_data)
    
    hourly_data = []
    for hour in hours:
        # Distribute traffic based on typical retail patterns
        multiplier = 1.8 if hour in peak_hours else 0.6
        estimated = int((total_in / len(hours)) * multiplier)
        hourly_data.append({
            "hour": f"{hour:02d}:00",
            "visitors": estimated,
            "is_peak": hour in peak_hours
        })
    
    return {
        "report_type": "hourly_traffic",
        "date_range": date_range,
        "total_stores": len(counter_data),
        "total_visitors": total_in,
        "peak_hours": [f"{h:02d}:00" for h in peak_hours],
        "hourly_data": hourly_data,
        "peak_hour_traffic": sum(d["visitors"] for d in hourly_data if d["is_peak"]),
        "off_peak_traffic": sum(d["visitors"] for d in hourly_data if not d["is_peak"])
    }

@api_router.get("/reports/advanced/weekday-comparison")
async def get_weekday_comparison(
    store_ids: Optional[str] = None,
    date_range: str = "1w",
    user: dict = Depends(require_auth)
):
    """Compare weekday vs weekend traffic"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    counter_data = await _fetch_live_counter_data(store_ids, allowed_stores)
    total_in = sum(s["total_in"] for s in counter_data)
    
    # Typical retail distribution: weekends have more traffic
    weekday_multiplier = 0.12  # Mon-Fri average
    weekend_multiplier = 0.20  # Sat-Sun average
    
    daily_data = [
        {"day": "Pazartesi", "visitors": int(total_in * 0.10), "type": "weekday"},
        {"day": "Salı", "visitors": int(total_in * 0.11), "type": "weekday"},
        {"day": "Çarşamba", "visitors": int(total_in * 0.12), "type": "weekday"},
        {"day": "Perşembe", "visitors": int(total_in * 0.13), "type": "weekday"},
        {"day": "Cuma", "visitors": int(total_in * 0.15), "type": "weekday"},
        {"day": "Cumartesi", "visitors": int(total_in * 0.21), "type": "weekend"},
        {"day": "Pazar", "visitors": int(total_in * 0.18), "type": "weekend"},
    ]
    
    weekday_total = sum(d["visitors"] for d in daily_data if d["type"] == "weekday")
    weekend_total = sum(d["visitors"] for d in daily_data if d["type"] == "weekend")
    
    return {
        "report_type": "weekday_comparison",
        "date_range": date_range,
        "total_stores": len(counter_data),
        "total_visitors": total_in,
        "daily_data": daily_data,
        "weekday_total": weekday_total,
        "weekend_total": weekend_total,
        "weekday_avg": weekday_total // 5,
        "weekend_avg": weekend_total // 2,
        "weekend_increase_percent": round((weekend_total / 2) / (weekday_total / 5) * 100 - 100, 1) if weekday_total > 0 else 0
    }

@api_router.get("/reports/advanced/store-comparison")
async def get_store_comparison(
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    user: dict = Depends(require_auth)
):
    """Compare performance between stores"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    counter_data = await _fetch_live_counter_data(store_ids, allowed_stores)
    
    if not counter_data:
        return {"report_type": "store_comparison", "stores": [], "message": "No stores found"}
    
    # Sort by visitors
    sorted_stores = sorted(counter_data, key=lambda x: x["current_visitors"], reverse=True)
    
    # Calculate averages
    avg_visitors = sum(s["current_visitors"] for s in counter_data) / len(counter_data)
    avg_occupancy = sum(s["occupancy_percent"] for s in counter_data) / len(counter_data)
    avg_in = sum(s["total_in"] for s in counter_data) / len(counter_data)
    
    store_comparison = []
    for store in sorted_stores:
        performance = "above" if store["current_visitors"] > avg_visitors else "below"
        deviation = round(((store["current_visitors"] - avg_visitors) / avg_visitors * 100) if avg_visitors > 0 else 0, 1)
        store_comparison.append({
            "store_id": store["store_id"],
            "store_name": store["store_name"],
            "city": store.get("city_name", ""),
            "district": store.get("district_name", ""),
            "current_visitors": store["current_visitors"],
            "total_in": store["total_in"],
            "total_out": store["total_out"],
            "occupancy_percent": store["occupancy_percent"],
            "capacity": store["capacity"],
            "status": store["status"],
            "performance": performance,
            "deviation_percent": deviation
        })
    
    return {
        "report_type": "store_comparison",
        "date_range": date_range,
        "total_stores": len(counter_data),
        "average_visitors": round(avg_visitors, 1),
        "average_occupancy": round(avg_occupancy, 1),
        "average_entries": round(avg_in, 1),
        "top_performer": store_comparison[0] if store_comparison else None,
        "bottom_performer": store_comparison[-1] if store_comparison else None,
        "stores": store_comparison
    }

@api_router.get("/reports/advanced/queue-analysis")
async def get_advanced_queue_analysis(
    store_ids: Optional[str] = None,
    date_range: str = "1d",
    user: dict = Depends(require_auth)
):
    """Advanced queue analysis with wait time estimation"""
    from permissions import get_user_allowed_stores
    allowed_stores = await get_user_allowed_stores(user)
    queue_data = await _fetch_live_queue_data(store_ids, allowed_stores)
    
    if not queue_data:
        return {"report_type": "queue_analysis", "stores": [], "message": "No queue data found"}
    
    # Calculate queue metrics
    total_queue = sum(s["total_queue_length"] for s in queue_data)
    avg_queue = total_queue / len(queue_data) if queue_data else 0
    
    # Estimate wait time (assuming 2 minutes per person in queue)
    avg_wait_time = avg_queue * 2  # minutes
    
    # Critical hours (typical busy hours)
    critical_hours = ["12:00-14:00", "17:00-19:00"]
    
    queue_analysis = []
    for store in queue_data:
        estimated_wait = store["total_queue_length"] * 2
        queue_analysis.append({
            "store_id": store["store_id"],
            "store_name": store["store_name"],
            "queue_length": store["total_queue_length"],
            "threshold": store["queue_threshold"],
            "status": store["status"],
            "estimated_wait_minutes": estimated_wait,
            "zones": store.get("zones", []),
            "exceeds_threshold": store["total_queue_length"] > store["queue_threshold"]
        })
    
    # Sort by queue length
    queue_analysis = sorted(queue_analysis, key=lambda x: x["queue_length"], reverse=True)
    
    return {
        "report_type": "queue_analysis",
        "date_range": date_range,
        "total_stores": len(queue_data),
        "total_queue_length": total_queue,
        "average_queue_length": round(avg_queue, 1),
        "average_wait_time_minutes": round(avg_wait_time, 1),
        "critical_hours": critical_hours,
        "stores_exceeding_threshold": len([s for s in queue_analysis if s["exceeds_threshold"]]),
        "stores": queue_analysis
    }

# ============== HISTORICAL DATA ENDPOINTS ==============

@api_router.get("/historical/counter")
async def get_historical_counter(
    store_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get historical counter data"""
    query = {}
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    data = await db.historical_counter.find(query, {"_id": 0}).sort([("date", -1), ("hour", -1)]).to_list(1000)
    return data

@api_router.get("/historical/queue")
async def get_historical_queue(
    store_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get historical queue data"""
    query = {}
    if store_id:
        query["store_id"] = store_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    data = await db.historical_queue.find(query, {"_id": 0}).sort([("date", -1), ("hour", -1)]).to_list(1000)
    return data

@api_router.get("/historical/analytics")
async def get_historical_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get historical analytics data"""
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    data = await db.historical_analytics.find(query, {"_id": 0}).sort([("date", -1), ("hour", -1)]).to_list(1000)
    return data

@api_router.get("/historical/summary")
async def get_historical_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Get summary of historical data"""
    # Default to last 7 days
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    
    query = {"date": {"$gte": start_date, "$lte": end_date}}
    
    counter_data = await db.historical_counter.find(query, {"_id": 0}).to_list(10000)
    queue_data = await db.historical_queue.find(query, {"_id": 0}).to_list(10000)
    analytics_data = await db.historical_analytics.find(query, {"_id": 0}).to_list(10000)
    
    # Aggregate by date
    daily_stats = {}
    for record in counter_data:
        date = record["date"]
        if date not in daily_stats:
            daily_stats[date] = {"date": date, "total_in": 0, "total_out": 0, "avg_visitors": 0, "count": 0}
        daily_stats[date]["total_in"] += record["total_in"]
        daily_stats[date]["total_out"] += record["total_out"]
        daily_stats[date]["avg_visitors"] += record["current_visitors"]
        daily_stats[date]["count"] += 1
    
    # Calculate averages
    for date, stats in daily_stats.items():
        if stats["count"] > 0:
            stats["avg_visitors"] = round(stats["avg_visitors"] / stats["count"], 1)
    
    return {
        "start_date": start_date,
        "end_date": end_date,
        "total_records": {
            "counter": len(counter_data),
            "queue": len(queue_data),
            "analytics": len(analytics_data)
        },
        "daily_stats": list(daily_stats.values()),
        "total_in": sum(r["total_in"] for r in counter_data),
        "total_out": sum(r["total_out"] for r in counter_data)
    }

@api_router.post("/historical/collect-now")
async def collect_historical_now(background_tasks: BackgroundTasks, admin: dict = Depends(require_admin)):
    """Manually trigger historical data collection"""
    background_tasks.add_task(collect_historical_data)
    return {"status": "started", "message": "Tarihsel veri toplama başlatıldı"}

# ============== PDF EXPORT ENDPOINT ==============

@api_router.get("/reports/export/pdf")
async def export_report_pdf(
    report_type: str = "counter",
    date_range: str = "1d",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_auth)
):
    """Export report as PDF with date range support"""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    # Styles with Turkish font support
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#3B82F6'), alignment=TA_CENTER, fontName=PDF_FONT_BOLD)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER, fontName=PDF_FONT)
    header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1F2937'), fontName=PDF_FONT_BOLD)
    
    # Date range label
    date_range_labels = {"1d": "1 Gün", "7d": "7 Gün", "1w": "7 Gün", "30d": "30 Gün", "1m": "30 Gün"}
    date_label = date_range_labels.get(date_range, date_range)
    
    # Title
    report_titles = {
        "counter": "Kişi Sayma Raporu",
        "queue": "Kuyruk Analizi Raporu",
        "demographics": "Demografik Analiz Raporu",
        "analytics": "Demografik Analiz Raporu",
        "store_comparison": "Mağaza Karşılaştırma Raporu"
    }
    elements.append(Paragraph(f"VMS360 - {report_titles.get(report_type, 'Rapor')}", title_style))
    elements.append(Paragraph(f"Tarih Aralığı: {date_label} | Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    if report_type == "counter":
        # Use get_counter_report for historical data support
        report = await get_counter_report(date_range=date_range, date_from=start_date, date_to=end_date)
        counter_data = report.get("stores", [])
        summary = report.get("summary", {})
        
        if counter_data:
            # Summary stats from report
            total_in = summary.get("total_in", 0)
            total_out = summary.get("total_out", 0)
            total_visitors = summary.get("current_visitors", 0)
            
            elements.append(Paragraph("Özet Bilgiler", header_style))
            summary_data = [
                ["Toplam Mağaza", "Anlık Ziyaretçi", "Toplam Giriş", "Toplam Çıkış"],
                [str(len(counter_data)), str(total_visitors), str(total_in), str(total_out)]
            ]
            summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
                ('GRID', (0, 0), (-1, -1), 1, colors.white),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*cm))
            
            # Store details
            elements.append(Paragraph("Mağaza Detayları", header_style))
            store_data = [["Mağaza", "Konum", "Anlık", "Giriş", "Çıkış", "Doluluk", "Durum"]]
            for store in counter_data[:20]:  # Limit to 20 stores
                status_tr = {"normal": "Normal", "warning": "Uyarı", "critical": "Kritik"}.get(store["status"], store["status"])
                store_data.append([
                    store["store_name"][:20],
                    f"{store.get('district_name', '')}",
                    str(store["current_visitors"]),
                    str(store["total_in"]),
                    str(store["total_out"]),
                    f"%{store['occupancy_percent']}",
                    status_tr
                ])
            
            store_table = Table(store_data, colWidths=[4*cm, 4*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
            store_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            elements.append(store_table)
    
    elif report_type == "queue":
        # Use the new meaningful queue report with date range
        queue_report = await get_queue_report(date_range=date_range)
        queue_data = queue_report.get("stores", [])
        summary = queue_report.get("summary", {})
        
        if queue_data:
            elements.append(Paragraph("Kuyruk Özeti", header_style))
            summary_data = [
                ["Toplam Mağaza", "Ort. Kuyruk", "Maks. Kuyruk", "Eşik Aşım Sayısı"],
                [
                    str(summary.get("total_stores", len(queue_data))), 
                    str(summary.get("avg_queue_overall", 0)),
                    str(summary.get("max_queue_observed", 0)),
                    str(summary.get("total_threshold_exceeds", 0))
                ]
            ]
            summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FEF3C7')),
                ('GRID', (0, 0), (-1, -1), 1, colors.white),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*cm))
            
            elements.append(Paragraph("Mağaza Kuyruk Durumu", header_style))
            queue_table_data = [["Mağaza", "Ort. Kuyruk", "Maks. Kuyruk", "Eşik Aşım", "Eşik", "Durum"]]
            for store in queue_data[:20]:
                status_tr = {"normal": "Normal", "warning": "Uyarı", "critical": "Kritik"}.get(store.get("status", ""), store.get("status", ""))
                queue_table_data.append([
                    store.get("store_name", "")[:25],
                    str(store.get("avg_queue_length", 0)),
                    str(store.get("max_queue_length", 0)),
                    str(store.get("threshold_exceed_count", 0)),
                    str(store.get("queue_threshold", 5)),
                    status_tr
                ])
            
            queue_table = Table(queue_table_data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm, 3.5*cm])
            queue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            elements.append(queue_table)
    
    elif report_type == "demographics" or report_type == "analytics":
        # Use get_analytics_report for historical data support
        analytics_report = await get_analytics_report(date_range=date_range)
        
        elements.append(Paragraph("Demografik Özet", header_style))
        gender = analytics_report.get("gender_distribution", {"Male": 0, "Female": 0})
        age = analytics_report.get("age_distribution", {})
        total = analytics_report.get("summary", {}).get("total_detections", 0)
        
        summary_data = [
            ["Toplam Tespit", "Erkek", "Kadın"],
            [str(total), str(gender.get("Male", 0)), str(gender.get("Female", 0))]
        ]
        summary_table = Table(summary_data, colWidths=[6*cm, 6*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EDE9FE')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("Yaş Dağılımı", header_style))
        age_data = [["Yaş Grubu", "Sayı", "Oran"]]
        for age_group, count in age.items():
            percent = round(count / total * 100, 1) if total > 0 else 0
            age_data.append([age_group, str(count), f"%{percent}"])
        
        age_table = Table(age_data, colWidths=[6*cm, 6*cm, 6*cm])
        age_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(age_table)
    
    elif report_type == "store_comparison":
        counter_data = await _fetch_live_counter_data(None, None)
        
        if counter_data:
            sorted_stores = sorted(counter_data, key=lambda x: x["current_visitors"], reverse=True)
            avg_visitors = sum(s["current_visitors"] for s in counter_data) / len(counter_data)
            
            elements.append(Paragraph("Mağaza Performans Karşılaştırması", header_style))
            elements.append(Paragraph(f"Ortalama Ziyaretçi: {round(avg_visitors, 1)}", subtitle_style))
            elements.append(Spacer(1, 0.3*cm))
            
            comp_data = [["Sıra", "Mağaza", "Ziyaretçi", "Giriş", "Doluluk", "Performans"]]
            for idx, store in enumerate(sorted_stores[:15], 1):
                deviation = round(((store["current_visitors"] - avg_visitors) / avg_visitors * 100) if avg_visitors > 0 else 0, 1)
                perf = f"+{deviation}%" if deviation >= 0 else f"{deviation}%"
                comp_data.append([
                    str(idx),
                    store["store_name"][:20],
                    str(store["current_visitors"]),
                    str(store["total_in"]),
                    f"%{store['occupancy_percent']}",
                    perf
                ])
            
            comp_table = Table(comp_data, colWidths=[1.5*cm, 6*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDF4')]),
            ]))
            elements.append(comp_table)
    
    # Footer
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER, fontName=PDF_FONT)
    elements.append(Paragraph("Bu rapor VMS360 Retail Panel tarafından otomatik olarak oluşturulmuştur.", footer_style))
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    filename = f"rapor_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============== SMTP SETTINGS ENDPOINTS ==============

@api_router.get("/settings/smtp")
async def get_smtp_settings(admin: dict = Depends(require_admin)):
    """Get SMTP settings (admin only)"""
    settings = await db.smtp_settings.find_one({}, {"_id": 0, "password": 0})
    return settings or {}

@api_router.post("/settings/smtp")
async def save_smtp_settings(settings: SMTPSettingsCreate, admin: dict = Depends(require_admin)):
    """Save SMTP settings (admin only)"""
    # Check if settings exist
    existing = await db.smtp_settings.find_one({})
    
    # If password is empty and existing settings exist, keep the old password
    password_to_use = settings.password
    if not password_to_use and existing and existing.get('password'):
        password_to_use = existing['password']
    
    smtp_data = SMTPSettings(
        host=settings.host,
        port=settings.port,
        username=settings.username,
        password=password_to_use,
        from_email=settings.from_email,
        from_name=settings.from_name,
        use_tls=settings.use_tls
    )
    
    doc = smtp_data.model_dump()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    if existing:
        doc['id'] = existing['id']
        await db.smtp_settings.replace_one({"id": existing['id']}, doc)
    else:
        await db.smtp_settings.insert_one(doc)
    
    # Return without password
    doc.pop('password', None)
    doc.pop('_id', None)
    return doc

@api_router.post("/settings/smtp/test")
async def test_smtp_settings(request: SMTPTestRequest, admin: dict = Depends(require_admin)):
    """Test SMTP settings by sending a test email"""
    settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="SMTP ayarları bulunamadı. Önce SMTP ayarlarını kaydedin.")
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = f"{settings['from_name']} <{settings['from_email']}>"
        msg['To'] = request.test_email
        msg['Subject'] = "VMS360 - SMTP Test Mesajı"
        
        body = """
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px;">
            <h2 style="color: #3B82F6;">VMS360 Retail Panel</h2>
            <p>Bu bir test mesajıdır. SMTP ayarlarınız başarıyla yapılandırıldı!</p>
            <hr style="border: 1px solid #E5E7EB; margin: 20px 0;">
            <p style="color: #6B7280; font-size: 12px;">
                Bu mesaj VMS360 Retail Panel tarafından gönderilmiştir.
            </p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))
        
        # Connect and send
        if settings.get('use_tls', True):
            server = smtplib.SMTP(settings['host'], settings['port'], timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        else:
            server = smtplib.SMTP_SSL(settings['host'], settings['port'], timeout=30)
            server.ehlo()
        
        server.login(settings['username'], settings['password'])
        server.send_message(msg)
        server.quit()
        
        return {"status": "success", "message": f"Test e-postası {request.test_email} adresine gönderildi."}
    except Exception as e:
        logger.error(f"SMTP test failed: {str(e)}")
        raise HTTPException(status_code=400, detail=f"E-posta gönderilemedi: {str(e)}")

# ============== SCHEDULED REPORTS ENDPOINTS ==============

@api_router.get("/scheduled-reports")
async def get_scheduled_reports(user: dict = Depends(require_auth)):
    """Get all scheduled reports"""
    reports = await db.scheduled_reports.find({}, {"_id": 0}).to_list(100)
    return reports

@api_router.post("/scheduled-reports")
@api_router.post("/scheduled-reports-v2")
async def create_scheduled_report(request: Request, user: dict = Depends(require_admin)):
    """Create a new scheduled report (admin only)"""
    # Get raw JSON body
    body = await request.json()
    logger.info(f"RAW REQUEST BODY: {body}")
    
    # Extract fields directly from body
    date_range = body.get('date_range', '1d')
    store_ids = body.get('store_ids', [])
    
    logger.info(f"Extracted: date_range={date_range}, store_ids={store_ids}")
    
    doc = {
        "id": str(uuid.uuid4()),
        "name": body.get('name'),
        "report_type": body.get('report_type'),
        "format": body.get('format', 'excel'),
        "frequency": body.get('frequency', 'daily'),
        "send_time": body.get('send_time', '08:00'),
        "send_day": body.get('send_day'),
        "recipients": body.get('recipients', []),
        "created_by": user['username'],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_active": True,
        "last_sent": None,
        # CRITICAL: Save filter fields directly from body
        "store_ids": store_ids,
        "date_range": date_range,
        "date_from": body.get('date_from'),
        "date_to": body.get('date_to'),
        "hour_from": body.get('hour_from'),
        "hour_to": body.get('hour_to'),
        "gender_filter": body.get('gender_filter'),
        "min_queue_length": body.get('min_queue_length', 0)
    }
    
    logger.info(f"FINAL DOC date_range={doc['date_range']}, store_ids={doc['store_ids']}")
    
    await db.scheduled_reports.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.put("/scheduled-reports/{report_id}")
async def update_scheduled_report(report_id: str, update: ScheduledReportUpdate, admin: dict = Depends(require_admin)):
    """Update a scheduled report (admin only)"""
    report = await db.scheduled_reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Planlı rapor bulunamadı")
    
    # Use exclude_unset=True to only include fields that were actually sent in the request
    # This prevents overwriting existing values with None defaults
    update_data = update.model_dump(exclude_unset=True)
    
    logger.info(f"Updating scheduled report {report_id}: {update_data}")
    
    if update_data:
        await db.scheduled_reports.update_one({"id": report_id}, {"$set": update_data})
    
    updated = await db.scheduled_reports.find_one({"id": report_id}, {"_id": 0})
    return updated

@api_router.delete("/scheduled-reports/{report_id}")
async def delete_scheduled_report(report_id: str, admin: dict = Depends(require_admin)):
    """Delete a scheduled report (admin only)"""
    result = await db.scheduled_reports.delete_one({"id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Planlı rapor bulunamadı")
    return {"status": "deleted"}

@api_router.post("/scheduled-reports/{report_id}/send-now")
async def send_report_now(report_id: str, background_tasks: BackgroundTasks, admin: dict = Depends(require_admin)):
    """Manually trigger sending a scheduled report"""
    report = await db.scheduled_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Planlı rapor bulunamadı")
    
    smtp_settings = await db.smtp_settings.find_one({}, {"_id": 0})
    if not smtp_settings:
        raise HTTPException(status_code=400, detail="SMTP ayarları yapılandırılmamış")
    
    background_tasks.add_task(send_scheduled_report, report, smtp_settings)
    return {"status": "queued", "message": "Rapor gönderimi başlatıldı"}

async def generate_report_data(report_type: str, filters: dict = None):
    """Generate report data based on type and filters"""
    from datetime import timedelta
    
    if filters is None:
        filters = {}
    
    store_ids = filters.get("store_ids", [])
    store_ids_str = ",".join(store_ids) if store_ids else None
    hour_from = filters.get("hour_from")
    hour_to = filters.get("hour_to")
    gender_filter = filters.get("gender_filter")
    min_queue_length = filters.get("min_queue_length", 0)
    date_range = filters.get("date_range", "1d")
    
    # Get store location info helper
    async def get_stores_with_location():
        stores = await db.stores.find({}, {"_id": 0}).to_list(500)
        districts = await db.districts.find({}, {"_id": 0}).to_list(500)
        cities = await db.cities.find({}, {"_id": 0}).to_list(500)
        regions = await db.regions.find({}, {"_id": 0}).to_list(500)
        
        districts_map = {d["id"]: d for d in districts}
        cities_map = {c["id"]: c for c in cities}
        regions_map = {r["id"]: r for r in regions}
        
        stores_map = {}
        for s in stores:
            district = districts_map.get(s.get("district_id"), {})
            city = cities_map.get(district.get("city_id"), {})
            region = regions_map.get(city.get("region_id"), {})
            stores_map[s["id"]] = {
                **s,
                "district_name": district.get("name", ""),
                "city_name": city.get("name", ""),
                "region_name": region.get("name", "")
            }
        return stores_map
    
    # ========== COUNTER REPORT (Kişi Sayma) ==========
    if report_type == "counter":
        report = await get_counter_report(
            store_ids=store_ids_str,
            date_range=date_range,
            allowed_stores=None
        )
        data = []
        for s in report.get("stores", []):
            data.append({
                "Mağaza": s.get("store_name", ""),
                "Bölge": s.get("region_name", ""),
                "İl": s.get("city_name", ""),
                "İlçe": s.get("district_name", ""),
                "Giriş": s.get("total_in", 0),
                "Çıkış": s.get("total_out", 0),
                "Mevcut": s.get("current_visitors", 0),
                "Kapasite": s.get("capacity", 0),
                "Doluluk %": s.get("occupancy_percent", 0),
                "Durum": s.get("status", "").upper()
            })
        # Add total row
        summary = report.get("summary", {})
        data.append({
            "Mağaza": "TOPLAM",
            "Bölge": "",
            "İl": "",
            "İlçe": "",
            "Giriş": summary.get("total_in", 0),
            "Çıkış": summary.get("total_out", 0),
            "Mevcut": summary.get("current_visitors", 0),
            "Kapasite": sum(s.get("capacity", 0) for s in report.get("stores", [])),
            "Doluluk %": summary.get("avg_occupancy", 0),
            "Durum": ""
        })
        return {"type": "Kişi Sayma Raporu", "data": data}
    
    # ========== QUEUE REPORT (Kuyruk Analizi) ==========
    elif report_type == "queue":
        report = await get_queue_report(
            store_ids=store_ids_str,
            date_range=date_range,
            allowed_stores=None
        )
        data = []
        for s in report.get("stores", []):
            data.append({
                "Mağaza": s.get("store_name", ""),
                "Bölge": s.get("region_name", ""),
                "İl": s.get("city_name", ""),
                "İlçe": s.get("district_name", ""),
                "Ort. Kuyruk": s.get("avg_queue_length", 0),  # MEANINGFUL: Average queue
                "Maks. Kuyruk": s.get("max_queue_length", 0),  # MEANINGFUL: Max queue
                "Eşik Aşım": s.get("threshold_exceed_count", 0),  # MEANINGFUL: Exceed count
                "Ölçüm Sayısı": s.get("measurement_count", 0),  # How many measurements
                "Eşik": s.get("queue_threshold", 0),
                "Durum": s.get("status", "").upper()
            })
        # Add total row with meaningful summary
        summary = report.get("summary", {})
        data.append({
            "Mağaza": "TOPLAM",
            "Bölge": "",
            "İl": "",
            "İlçe": "",
            "Ort. Kuyruk": summary.get("avg_queue_overall", 0),
            "Maks. Kuyruk": summary.get("max_queue_observed", 0),
            "Eşik Aşım": summary.get("total_threshold_exceeds", 0),
            "Ölçüm Sayısı": "",
            "Eşik": "",
            "Durum": ""
        })
        return {"type": "Kuyruk Analizi Raporu", "data": data}
    
    # ========== ANALYTICS REPORT (Yaş/Cinsiyet) ==========
    elif report_type == "analytics":
        logger.info(f"Generating analytics report with filters: store_ids={store_ids}, date_range={date_range}")
        
        report = await get_analytics_report(
            store_ids=store_ids_str,
            date_range=date_range
        )
        
        data = []
        
        # Add data source info as first row for transparency
        data_source = report.get("data_source", "unknown")
        data_source_note = report.get("data_source_note", "")
        if data_source == "live_vms":
            data.append({
                "Mağaza": f"⚠️ VERİ KAYNAĞI: {data_source_note}",
                "Bölge": "",
                "İl": "",
                "İlçe": "",
                "Tespit": "",
                "Erkek": "",
                "Kadın": "",
                "Erkek %": "",
                "Kadın %": ""
            })
        
        for s in report.get("stores", []):
            data.append({
                "Mağaza": s.get("store_name", ""),
                "Bölge": s.get("region_name", ""),
                "İl": s.get("city_name", ""),
                "İlçe": s.get("district_name", ""),
                "Tespit": s.get("total_detections", 0),
                "Erkek": s.get("male_count", 0),
                "Kadın": s.get("female_count", 0),
                "Erkek %": s.get("male_percent", 0),
                "Kadın %": s.get("female_percent", 0)
            })
        
        # Add total row
        summary = report.get("summary", {})
        data.append({
            "Mağaza": "TOPLAM",
            "Bölge": "",
            "İl": "",
            "İlçe": "",
            "Tespit": summary.get("total_detections", 0),
            "Erkek": summary.get("male_count", 0),
            "Kadın": summary.get("female_count", 0),
            "Erkek %": summary.get("male_percent", 0),
            "Kadın %": summary.get("female_percent", 0)
        })
        
        logger.info(f"Analytics report generated: {len(data)} rows, data_source={data_source}")
        return {"type": "Yaş/Cinsiyet Analizi", "data": data, "data_source": data_source}
    
    # ========== HOURLY TRAFFIC (Saatlik Trafik - Gerçek Veri) ==========
    elif report_type == "hourly_traffic":
        now = datetime.now(timezone.utc)
        today = now.strftime("%Y-%m-%d")
        
        # Build query
        query = {"date": today}
        if store_ids:
            query["store_id"] = {"$in": store_ids}
        
        # Get hourly data from counter_snapshots - aggregate by hour
        pipeline = [
            {"$match": query},
            {"$group": {
                "_id": "$hour",
                "total_in": {"$max": "$total_in"},  # Get max (cumulative) for each hour
                "store_count": {"$sum": 1}
            }},
            {"$sort": {"_id": 1}}
        ]
        hourly_results = await db.counter_snapshots.aggregate(pipeline).to_list(24)
        
        # Calculate hourly increments (difference between consecutive hours)
        hourly_map = {r["_id"]: r["total_in"] for r in hourly_results}
        
        start_hour = hour_from if hour_from is not None else 8
        end_hour = hour_to if hour_to is not None else 22
        peak_hours = [11, 12, 13, 17, 18, 19]
        
        hourly_data = []
        prev_total = 0
        for hour in range(start_hour, end_hour + 1):
            current_total = hourly_map.get(hour, prev_total)
            visitors = max(0, current_total - prev_total) if prev_total > 0 else current_total
            prev_total = current_total
            
            hourly_data.append({
                "Saat": f"{hour:02d}:00",
                "Ziyaretçi": visitors,
                "Yoğun Saat": "Evet" if hour in peak_hours else "Hayır"
            })
        
        # Add total row
        total_visitors = sum(h["Ziyaretçi"] for h in hourly_data)
        hourly_data.append({
            "Saat": "TOPLAM",
            "Ziyaretçi": total_visitors,
            "Yoğun Saat": ""
        })
        
        return {"type": "Saatlik Trafik", "data": hourly_data}
    
    # ========== WEEKDAY COMPARISON (Haftalık Karşılaştırma - Gerçek Veri) ==========
    elif report_type == "weekday_comparison":
        now = datetime.now(timezone.utc)
        
        # Get last 7 days data
        daily_data = []
        day_names_tr = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        
        for i in range(7):
            day_date = now - timedelta(days=6-i)
            date_str = day_date.strftime("%Y-%m-%d")
            day_name = day_names_tr[day_date.weekday()]
            
            # Query for this day's data
            query = {"date": date_str}
            if store_ids:
                query["store_id"] = {"$in": store_ids}
            
            # Get max total_in for each store on this day (last snapshot of the day)
            pipeline = [
                {"$match": query},
                {"$sort": {"hour": -1, "minute": -1}},
                {"$group": {
                    "_id": "$store_id",
                    "total_in": {"$first": "$total_in"}
                }}
            ]
            day_results = await db.counter_snapshots.aggregate(pipeline).to_list(100)
            day_total = sum(r["total_in"] for r in day_results) if day_results else 0
            
            daily_data.append({
                "Gün": day_name,
                "Tarih": date_str,
                "Ziyaretçi": day_total
            })
        
        # Add total row
        total_visitors = sum(d["Ziyaretçi"] for d in daily_data)
        daily_data.append({
            "Gün": "TOPLAM",
            "Tarih": "",
            "Ziyaretçi": total_visitors
        })
        
        return {"type": "Haftalık Karşılaştırma", "data": daily_data}
    
    # ========== STORE COMPARISON (Mağaza Karşılaştırma) ==========
    elif report_type == "store_comparison":
        stores_map = await get_stores_with_location()
        counter_data = await _fetch_live_counter_data(store_ids_str, None)
        
        store_data = []
        for s in counter_data:
            store_info = stores_map.get(s["store_id"], {})
            store_data.append({
                "Mağaza": s["store_name"],
                "Bölge": store_info.get("region_name", ""),
                "İl": store_info.get("city_name", ""),
                "İlçe": store_info.get("district_name", ""),
                "Toplam Giriş": s["total_in"],
                "Toplam Çıkış": s["total_out"],
                "Mevcut Ziyaretçi": s["current_visitors"],
                "Doluluk %": s.get("occupancy_percent", 0)
            })
        
        # Add total row
        total_in = sum(s["Toplam Giriş"] for s in store_data)
        total_out = sum(s["Toplam Çıkış"] for s in store_data)
        total_current = sum(s["Mevcut Ziyaretçi"] for s in store_data)
        avg_occupancy = sum(s["Doluluk %"] for s in store_data) / len(store_data) if store_data else 0
        
        store_data.append({
            "Mağaza": "TOPLAM",
            "Bölge": "",
            "İl": "",
            "İlçe": "",
            "Toplam Giriş": total_in,
            "Toplam Çıkış": total_out,
            "Mevcut Ziyaretçi": total_current,
            "Doluluk %": round(avg_occupancy, 1)
        })
        
        return {"type": "Mağaza Karşılaştırma", "data": store_data}
    
    # ========== QUEUE ANALYSIS (Kuyruk Analizi - Eski format) ==========
    elif report_type == "queue_analysis":
        stores_map = await get_stores_with_location()
        queue_data = await _fetch_live_queue_data(store_ids_str, None)
        
        queue_list = []
        for s in queue_data:
            if s["total_queue_length"] >= min_queue_length:
                store_info = stores_map.get(s["store_id"], {})
                queue_list.append({
                    "Mağaza": s["store_name"],
                    "Bölge": store_info.get("region_name", ""),
                    "İl": store_info.get("city_name", ""),
                    "İlçe": store_info.get("district_name", ""),
                    "Kuyruk Uzunluğu": s["total_queue_length"],
                    "Tahmini Bekleme (dk)": s["total_queue_length"] * 2,
                    "Durum": "Normal" if s["total_queue_length"] < 5 else "Yoğun"
                })
        
        # Add total row
        total_queue = sum(q["Kuyruk Uzunluğu"] for q in queue_list)
        queue_list.append({
            "Mağaza": "TOPLAM",
            "Bölge": "",
            "İl": "",
            "İlçe": "",
            "Kuyruk Uzunluğu": total_queue,
            "Tahmini Bekleme (dk)": total_queue * 2,
            "Durum": ""
        })
        
        return {"type": "Kuyruk Analizi", "data": queue_list}
    
    # ========== DEMOGRAPHICS (Demografik Analiz) ==========
    elif report_type == "demographics":
        # Get camera IDs for selected stores
        allowed_camera_ids = None
        if store_ids:
            stores = await db.stores.find({"id": {"$in": store_ids}}, {"_id": 0}).to_list(100)
            camera_ids = []
            for store in stores:
                analytics_ids = store.get("analytics_camera_ids", [])
                if store.get("analytics_camera_id") and store["analytics_camera_id"] not in analytics_ids:
                    analytics_ids.append(store["analytics_camera_id"])
                camera_ids.extend(analytics_ids)
            if camera_ids:
                allowed_camera_ids = camera_ids
        
        analytics = await _fetch_analytics_data(store_ids=store_ids_str, allowed_camera_ids=allowed_camera_ids)
        gender = analytics.get("gender_distribution", {})
        age = analytics.get("age_distribution", {})
        total = analytics.get("total_events", 0)
        
        demo_data = []
        for g, count in gender.items():
            if g in ["Male", "Female"]:
                if gender_filter and gender_filter != g:
                    continue
                demo_data.append({
                    "Kategori": "Cinsiyet",
                    "Değer": "Erkek" if g == "Male" else "Kadın",
                    "Sayı": count,
                    "Yüzde %": round(count / total * 100, 1) if total > 0 else 0
                })
        
        if not gender_filter:
            for a, count in age.items():
                demo_data.append({
                    "Kategori": "Yaş Grubu",
                    "Değer": a,
                    "Sayı": count,
                    "Yüzde %": round(count / total * 100, 1) if total > 0 else 0
                })
        
        # Add total row
        total_count = sum(d["Sayı"] for d in demo_data if d["Kategori"] == "Cinsiyet")
        demo_data.append({
            "Kategori": "TOPLAM",
            "Değer": "",
            "Sayı": total_count,
            "Yüzde %": 100.0 if total_count > 0 else 0
        })
        
        return {"type": "Demografik Analiz", "data": demo_data}
    
    # ========== ALL REPORTS (Tüm Raporlar) ==========
    else:  # all
        all_data = []
        
        # Kişi Sayma
        counter = await generate_report_data("counter", filters)
        for item in counter.get("data", []):
            item["Rapor Tipi"] = "Kişi Sayma"
            all_data.append(item)
        
        # Kuyruk
        queue = await generate_report_data("queue", filters)
        for item in queue.get("data", []):
            item["Rapor Tipi"] = "Kuyruk Analizi"
            all_data.append(item)
        
        # Yaş/Cinsiyet
        analytics = await generate_report_data("analytics", filters)
        for item in analytics.get("data", []):
            item["Rapor Tipi"] = "Yaş/Cinsiyet"
            all_data.append(item)
        
        return {"type": "Tüm Raporlar", "data": all_data}

async def send_scheduled_report(report: dict, smtp_settings: dict):
    """Send scheduled report via email"""
    try:
        logger.info(f"Generating scheduled report: {report.get('name')} - Type: {report.get('report_type')}")
        logger.info(f"Report raw date_range value: {report.get('date_range')} (type: {type(report.get('date_range'))})")
        
        # Build filters from report settings
        # CRITICAL FIX: Use 'or' instead of default value because None is a valid dict value
        # dict.get("key", "default") returns None if key exists with None value
        # Using 'or' ensures we get "1d" when value is None OR missing
        date_range_value = report.get("date_range") or "1d"
        
        filters = {
            "store_ids": report.get("store_ids") or [],
            "date_range": date_range_value,
            "hour_from": report.get("hour_from"),
            "hour_to": report.get("hour_to"),
            "gender_filter": report.get("gender_filter"),
            "min_queue_length": report.get("min_queue_length") or 0
        }
        
        logger.info(f"Report filters (after fix): {filters}")
        
        report_data = await generate_report_data(report["report_type"], filters)
        
        logger.info(f"Report data generated: {report_data.get('type')} - {len(report_data.get('data', []))} items")
        
        # Create email
        msg = MIMEMultipart()
        msg['From'] = f"{smtp_settings['from_name']} <{smtp_settings['from_email']}>"
        msg['To'] = ", ".join(report['recipients'])
        msg['Subject'] = f"VMS360 - {report['name']} ({datetime.now().strftime('%d.%m.%Y')})"
        
        # Email body
        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; padding: 20px; background: #f9fafb;">
            <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px;">
                <h2 style="color: #3B82F6; margin-bottom: 20px;">VMS360 Retail Panel</h2>
                <h3 style="color: #1F2937;">{report['name']}</h3>
                <p style="color: #6B7280;">Rapor Tipi: <strong>{report_data['type']}</strong></p>
                <p style="color: #6B7280;">Tarih: <strong>{datetime.now().strftime('%d.%m.%Y %H:%M')}</strong></p>
                <hr style="border: 1px solid #E5E7EB; margin: 20px 0;">
                <p style="color: #6B7280;">Rapor dosyası ekte yer almaktadır.</p>
                <p style="color: #9CA3AF; font-size: 12px; margin-top: 30px;">
                    Bu e-posta VMS360 Retail Panel planlı rapor sistemi tarafından otomatik olarak gönderilmiştir.
                </p>
            </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))
        
        # Generate attachment
        format_type = report.get('format', 'json')
        if format_type == 'json':
            attachment_data = json.dumps(report_data['data'], indent=2, ensure_ascii=False).encode('utf-8')
            filename = f"rapor_{report['report_type']}_{datetime.now().strftime('%Y%m%d')}.json"
            mime_type = "application/json"
        elif format_type == 'csv':
            import csv
            from io import StringIO
            output = StringIO()
            data = report_data['data']
            if isinstance(data, list) and len(data) > 0:
                writer = csv.DictWriter(output, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            attachment_data = output.getvalue().encode('utf-8')
            filename = f"rapor_{report['report_type']}_{datetime.now().strftime('%Y%m%d')}.csv"
            mime_type = "text/csv"
        else:  # excel
            import xlsxwriter
            output = BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            
            # Türkçe stil formatları
            header_format = workbook.add_format({
                'bold': True, 
                'bg_color': '#3B82F6', 
                'font_color': 'white', 
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            cell_format = workbook.add_format({'border': 1, 'align': 'left'})
            number_format = workbook.add_format({'border': 1, 'num_format': '#,##0', 'align': 'right'})
            percent_format = workbook.add_format({'border': 1, 'num_format': '0.0"%"', 'align': 'right'})
            total_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'align': 'left'})
            total_number_format = workbook.add_format({'bold': True, 'bg_color': '#FEF3C7', 'border': 1, 'num_format': '#,##0', 'align': 'right'})
            
            worksheet = workbook.add_worksheet("Rapor")
            data = report_data['data']
            
            logger.info(f"Excel generation - Data count: {len(data) if data else 0}")
            
            if isinstance(data, list) and len(data) > 0:
                headers = list(data[0].keys())
                logger.info(f"Excel generation - Headers: {headers}")
                
                # Header row
                for col, header in enumerate(headers):
                    worksheet.write(0, col, header, header_format)
                    worksheet.set_column(col, col, 18)  # Set column width
                
                # Data rows
                for row, item in enumerate(data, 1):
                    is_total = item.get("Mağaza") == "TOPLAM" or item.get("Saat") == "TOPLAM" or item.get("Gün") == "TOPLAM"
                    for col, key in enumerate(headers):
                        value = item.get(key, "")
                        if is_total:
                            if isinstance(value, (int, float)):
                                worksheet.write(row, col, value, total_number_format)
                            else:
                                worksheet.write(row, col, str(value) if value else "", total_format)
                        elif isinstance(value, (int, float)) and "%" in key:
                            worksheet.write(row, col, value, percent_format)
                        elif isinstance(value, (int, float)):
                            worksheet.write(row, col, value, number_format)
                        else:
                            worksheet.write(row, col, str(value) if value else "", cell_format)
                
                logger.info(f"Excel generation - Wrote {len(data)} rows")
            else:
                worksheet.write(0, 0, "Veri bulunamadı", cell_format)
                logger.warning("Excel generation - No data to write")
            
            workbook.close()
            output.seek(0)
            attachment_data = output.read()
            logger.info(f"Excel generation - File size: {len(attachment_data)} bytes")
            filename = f"rapor_{report['report_type']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        # Attach file
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
        # Send email
        if smtp_settings.get('use_tls', True):
            server = smtplib.SMTP(smtp_settings['host'], smtp_settings['port'])
            server.starttls()
        else:
            server = smtplib.SMTP_SSL(smtp_settings['host'], smtp_settings['port'])
        
        server.login(smtp_settings['username'], smtp_settings['password'])
        server.send_message(msg)
        server.quit()
        
        # Update last_sent
        await db.scheduled_reports.update_one(
            {"id": report["id"]},
            {"$set": {"last_sent": datetime.now(timezone.utc).isoformat()}}
        )
        
        logger.info(f"Scheduled report '{report['name']}' sent to {report['recipients']}")
    except Exception as e:
        logger.error(f"Failed to send scheduled report: {str(e)}")


# ============== ADVANCED ANALYTICS EXPORT ENDPOINTS ==============

from pydantic import BaseModel

class AnalyticsExportData(BaseModel):
    report_type: str
    tab: str
    period: str
    filters: dict
    data: dict

@api_router.post("/reports/export/excel")
async def export_analytics_excel(
    export_data: AnalyticsExportData,
    user: dict = Depends(require_auth)
):
    """Export advanced analytics to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Gelişmiş Analitik"
    
    # Styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "VMS360 - Gelişmiş Analitik Raporu"
    ws['A1'].font = Font(size=16, bold=True, color="3B82F6")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A3'] = f"Filtreler: Bölge: {export_data.filters.get('region', 'Tümü')}, Şehir: {export_data.filters.get('city', 'Tümü')}, Mağaza: {export_data.filters.get('store', 'Tümü')}"
    
    row = 5
    data = export_data.data
    
    # Summary section
    if data.get('summary'):
        summary = data['summary']
        ws[f'A{row}'] = "ÖZET"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        quick_stats = summary.get('quick_stats', {})
        stats = [
            ("Bugünkü Ziyaretçi", quick_stats.get('today_visitors', 0)),
            ("Ortalama Doluluk", f"{quick_stats.get('avg_occupancy', 0)}%"),
            ("Ort. Bekleme Süresi", f"{quick_stats.get('avg_wait_time_min', 0)} dk"),
            ("Toplam Mağaza", quick_stats.get('total_stores', 0))
        ]
        
        for i, (label, value) in enumerate(stats):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        row += 1
    
    # Hourly traffic
    if data.get('hourly_traffic'):
        hourly = data['hourly_traffic']
        ws[f'A{row}'] = "SAATLİK TRAFİK"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Saat", "Giriş", "Çıkış"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row += 1
        
        for hour_data in hourly.get('hourly_data', []):
            ws.cell(row=row, column=1, value=hour_data.get('hour_label', '')).border = thin_border
            ws.cell(row=row, column=2, value=hour_data.get('in_count', 0)).border = thin_border
            ws.cell(row=row, column=3, value=hour_data.get('out_count', 0)).border = thin_border
            row += 1
        row += 1
    
    # Demographics
    if data.get('demographics'):
        demo = data['demographics']
        ws[f'A{row}'] = "DEMOGRAFİ"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Gender data
        if demo.get('gender_data'):
            ws[f'A{row}'] = "Cinsiyet Dağılımı"
            row += 1
            for g in demo['gender_data']:
                ws[f'A{row}'] = "Erkek" if g.get('gender') == 'Male' else 'Kadın'
                ws[f'B{row}'] = g.get('count', 0)
                row += 1
        row += 1
        
        # Age data
        if demo.get('age_data'):
            ws[f'A{row}'] = "Yaş Dağılımı"
            row += 1
            for a in demo['age_data']:
                ws[f'A{row}'] = a.get('age_group', '')
                ws[f'B{row}'] = a.get('count', 0)
                row += 1
        row += 1
    
    # Store comparison
    if data.get('store_comparison'):
        stores = data['store_comparison']
        ws[f'A{row}'] = "MAĞAZA KARŞILAŞTIRMASI"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Mağaza", "Ziyaretçi/Gün", "Doluluk %", "Dönüşüm %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row += 1
        
        for store in stores.get('stores', []):
            ws.cell(row=row, column=1, value=store.get('store_name', '')).border = thin_border
            ws.cell(row=row, column=2, value=store.get('visitors_per_day', 0)).border = thin_border
            ws.cell(row=row, column=3, value=store.get('occupancy_percent', 0)).border = thin_border
            ws.cell(row=row, column=4, value=store.get('conversion_rate', 0)).border = thin_border
            row += 1
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gelismis_analitik_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@api_router.post("/reports/export/pdf")
async def export_analytics_pdf(
    export_data: AnalyticsExportData,
    user: dict = Depends(require_auth)
):
    """Export advanced analytics to PDF"""
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    # Styles - Use registered fonts
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#3B82F6'), alignment=TA_CENTER, fontName=PDF_FONT_BOLD)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER, fontName=PDF_FONT)
    header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#1F2937'), fontName=PDF_FONT_BOLD)
    
    # Title
    elements.append(Paragraph("VMS360 - Gelişmiş Analitik Raporu", title_style))
    elements.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    filters_text = f"Filtreler: Bölge: {export_data.filters.get('region', 'Tümü')}, Şehir: {export_data.filters.get('city', 'Tümü')}, Mağaza: {export_data.filters.get('store', 'Tümü')}"
    elements.append(Paragraph(filters_text, subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    data = export_data.data
    
    # Summary
    if data.get('summary'):
        summary = data['summary']
        quick_stats = summary.get('quick_stats', {})
        
        elements.append(Paragraph("Özet Bilgiler", header_style))
        summary_data = [
            ["Bugünkü Ziyaretçi", "Ortalama Doluluk", "Ort. Bekleme Süresi", "Toplam Mağaza"],
            [
                str(quick_stats.get('today_visitors', 0)),
                f"{quick_stats.get('avg_occupancy', 0)}%",
                f"{quick_stats.get('avg_wait_time_min', 0)} dk",
                str(quick_stats.get('total_stores', 0))
            ]
        ]
        summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5*cm))
    
    # Store comparison
    if data.get('store_comparison') and data['store_comparison'].get('stores'):
        elements.append(Paragraph("Mağaza Performans Karşılaştırması", header_style))
        store_data = [["Mağaza", "Ziyaretçi/Gün", "Doluluk %", "Dönüşüm %"]]
        for store in data['store_comparison']['stores'][:10]:  # Top 10
            store_data.append([
                store.get('store_name', ''),
                str(store.get('visitors_per_day', 0)),
                f"{store.get('occupancy_percent', 0)}%",
                f"{store.get('conversion_rate', 0)}%"
            ])
        
        store_table = Table(store_data, colWidths=[7*cm, 4*cm, 4*cm, 4*cm])
        store_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        elements.append(store_table)
        elements.append(Spacer(1, 0.5*cm))
    
    # Demographics
    if data.get('demographics'):
        demo = data['demographics']
        elements.append(Paragraph("Demografik Analiz", header_style))
        
        if demo.get('gender_data'):
            gender_data = [["Cinsiyet", "Sayı", "Oran"]]
            total = sum(g.get('count', 0) for g in demo['gender_data'])
            for g in demo['gender_data']:
                count = g.get('count', 0)
                pct = round(count / total * 100, 1) if total > 0 else 0
                gender_data.append([
                    "Erkek" if g.get('gender') == 'Male' else 'Kadın',
                    str(count),
                    f"%{pct}"
                ])
            
            gender_table = Table(gender_data, colWidths=[6*cm, 5*cm, 5*cm])
            gender_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),  # Turkish font for data rows
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ]))
            elements.append(gender_table)
            elements.append(Spacer(1, 0.3*cm))
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=gelismis_analitik_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()

